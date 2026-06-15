from __future__ import annotations

from copy import copy
from html.parser import HTMLParser
from pathlib import Path
from datetime import datetime

import nbformat
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "docs" / "임시_stage2_sinusoidal_added_20260614_132150.xlsx"
FALLBACK_WORKBOOK_PATH = ROOT / "docs" / "임시_stage2_sinusoidal_added_rowwise.xlsx"
NOTEBOOK_PATH = ROOT / "notebooks" / "sasrec_timeaware_bpi2012_colab_train_09_260612.ipynb"
SOURCE_SHEET_NAME = "2-2.Attnbias_anchor"
TARGET_SHEET_NAME = "2-3.Sinusoidal"


DETAIL_METRICS = [
    "best_valid_full_ndcg@10",
    "best_valid_full_hr@10",
    "best_valid_full_ndcg@5",
    "best_valid_full_hr@5",
    "best_valid_full_mrr",
    "best_test_at_best_valid_full_ndcg@10",
    "best_test_at_best_valid_full_hr@10",
    "best_test_at_best_valid_full_ndcg@5",
    "best_test_at_best_valid_full_hr@5",
    "best_test_at_best_valid_full_mrr",
    "best_valid_sampled_ndcg@10",
    "best_valid_sampled_hr@10",
    "best_valid_sampled_ndcg@5",
    "best_valid_sampled_hr@5",
    "best_valid_sampled_mrr",
    "best_test_at_best_valid_sampled_ndcg@10",
    "best_test_at_best_valid_sampled_hr@10",
    "best_test_at_best_valid_sampled_ndcg@5",
    "best_test_at_best_valid_sampled_hr@5",
    "best_test_at_best_valid_sampled_mrr",
    "best_valid_task_accuracy",
    "best_valid_task_macro_f1",
    "best_valid_task_top5_accuracy",
    "best_valid_task_top10_accuracy",
    "best_test_at_best_valid_task_accuracy",
    "best_test_at_best_valid_task_macro_f1",
    "best_test_at_best_valid_task_top5_accuracy",
    "best_test_at_best_valid_task_top10_accuracy",
]


SUMMARY_VARIANTS = [
    "anchor_baseline",
    "anchor_bucket_b9",
    "anchor_attnbias_dstart_b9",
    "anchor_sinusoidal_dprev",
    "anchor_sinusoidal_dstart",
    "refine_baseline",
    "refine_bucket_b9",
    "refine_dstart_bucket_b9",
    "refine_dstart_continuous",
    "refine_attnbias_dstart_b9",
    "refine_sinusoidal_dprev",
    "refine_sinusoidal_dstart",
]


SUMMARY_METRICS = [
    "best_valid_full_ndcg@10",
    "best_test_at_best_valid_full_ndcg@10",
    "best_valid_full_ndcg@5",
    "best_test_at_best_valid_full_ndcg@5",
]


class SimpleTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._in_cell = False
        self._buf = ""
        self._current_row: list[str] = []

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self._current_row = []
        elif tag in ("th", "td"):
            self._in_cell = True
            self._buf = ""

    def handle_endtag(self, tag):
        if tag in ("th", "td") and self._in_cell:
            self._current_row.append(self._buf.strip())
            self._in_cell = False
        elif tag == "tr" and self._current_row:
            self.rows.append(self._current_row)

    def handle_data(self, data):
        if self._in_cell:
            self._buf += data


def parse_run_table() -> pd.DataFrame:
    nb = nbformat.read(NOTEBOOK_PATH, as_version=4)
    html = nb.cells[44]["outputs"][0]["data"]["text/html"]

    parser = SimpleTableParser()
    parser.feed(html)

    header = parser.rows[0][1:]
    records = []
    for row in parser.rows[1:]:
        values = row[1:]
        if len(values) != len(header):
            continue
        records.append(dict(zip(header, values)))

    df = pd.DataFrame(records)
    return df.apply(lambda col: col.map(convert_value))


def convert_value(value):
    if value in ("", "None", "nan", "NaN"):
        return None
    if isinstance(value, str):
        lower = value.lower()
        if lower == "true":
            return True
        if lower == "false":
            return False
        try:
            if "." not in value and "e" not in lower:
                return int(value)
            return float(value)
        except ValueError:
            return value
    return value


def copy_style(src, dst) -> None:
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def write_cell(ws, row: int, col: int, value, style_cell=None):
    cell = ws.cell(row=row, column=col, value=value)
    if style_cell is not None:
        copy_style(style_cell, cell)
    return cell


def set_column_widths(ws):
    widths = {
        "B": 6,
        "C": 12,
        "D": 24,
        "E": 34,
        "F": 8,
        "G": 10,
        "H": 14,
        "I": 14,
        "J": 18,
        "K": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for col_idx in range(12, 40):
        ws.column_dimensions[column_letter(col_idx)].width = 12


def column_letter(idx: int) -> str:
    result = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def main() -> None:
    run_df = parse_run_table()
    summary_df = (
        run_df.groupby("variant")[DETAIL_METRICS]
        .agg(["mean", "std"])
        .reindex(SUMMARY_VARIANTS)
    )

    wb = load_workbook(WORKBOOK_PATH)
    src = wb[SOURCE_SHEET_NAME]
    if TARGET_SHEET_NAME in wb.sheetnames:
        del wb[TARGET_SHEET_NAME]
    ws = wb.create_sheet(TARGET_SHEET_NAME)

    ws.freeze_panes = "A14"
    set_column_widths(ws)

    # Title / summary styles from source sheet.
    title_style = src["B2"]
    text_style = src["B4"]
    label_style = src["C6"]
    small_header_style = src["I8"]
    group_header_style = src["I14"]
    detail_header_style = src["B15"]
    section_title_style = src["B41"]
    meanstd_group_style = src["I42"]
    meanstd_subheader_style = src["I43"]
    meanstd_header_style = src["B44"]

    write_cell(ws, 2, 2, "SUMMARY", title_style)

    write_cell(ws, 4, 2, " anchor/refine baseline + time-aware (sinusoidal, attention bias, bucket, continuous)", text_style)
    write_cell(ws, 4, 9, "main metric(full ranking, NDCG@10) 기준", text_style)
    write_cell(ws, 5, 9, "  overall best는 anchor baseline(test mean 0.8898)", text_style)
    write_cell(ws, 6, 3, "sinusoidal", label_style)
    write_cell(ws, 6, 4, "log1p(delta)에 대해 sin/cos encoding을 적용하는 additive 방식", text_style)
    write_cell(ws, 6, 9, "  refine 기준 best time-aware는 sinusoidal_dstart(test mean 0.8644)", text_style)
    write_cell(ws, 7, 3, "delta_prev", label_style)
    write_cell(ws, 7, 4, "직전 이벤트와의 시간 차(초 단위)를 시간 정보로 사용", text_style)
    write_cell(ws, 8, 3, "delta_start", label_style)
    write_cell(ws, 8, 4, "첫 이벤트부터의 누적 시간(초 단위)를 시간 정보로 사용", text_style)
    write_cell(ws, 10, 3, "비교 대상", label_style)
    write_cell(ws, 10, 4, "anchor/refine baseline, bucket, attention bias, continuous, sinusoidal", text_style)

    # Summary mini-table
    summary_headers = ["variant", "split", "metric", "mean", "std"]
    for offset, header in enumerate(summary_headers, start=9):
        write_cell(ws, 8, offset, header, small_header_style)

    summary_row = 9
    for variant in SUMMARY_VARIANTS:
        for metric in SUMMARY_METRICS:
            split = "valid" if metric.startswith("best_valid_") else "test"
            metric_name = metric.replace("best_valid_", "").replace("best_test_at_best_valid_", "")
            mean_value = summary_df.loc[variant, (metric, "mean")]
            std_value = summary_df.loc[variant, (metric, "std")]
            row_values = [variant, split, metric_name, mean_value, std_value]
            for col, value in enumerate(row_values, start=9):
                write_cell(ws, summary_row, col, value, small_header_style if summary_row == 8 else text_style)
            summary_row += 1

    detail_start = summary_row + 2

    # Detailed run table
    detail_group_row = detail_start
    detail_header_row = detail_start + 1
    detail_data_start = detail_start + 2

    ws.merge_cells(start_row=detail_group_row, start_column=9, end_row=detail_group_row, end_column=17)
    ws.merge_cells(start_row=detail_group_row, start_column=18, end_row=detail_group_row, end_column=26)
    write_cell(ws, detail_group_row, 9, "VALID set 성능", group_header_style)
    write_cell(ws, detail_group_row, 18, "TEST set 성능", group_header_style)

    detail_headers = [
        "SEQ",
        "Dataset",
        "variant",
        "run_name",
        "seed",
        "maxlen",
        "dropout_rate",
        "time_encoding",
        "평가 방식",
        "best epoch 기준",
        "NDCG@10",
        "Hit@10",
        "NDCG@5",
        "Hit@5",
        "MRR",
        "accuracy",
        "macro_f1",
        "top5_acc",
        "top10_acc",
        "NDCG@10",
        "Hit@10",
        "NDCG@5",
        "Hit@5",
        "MRR",
        "accuracy",
        "macro_f1",
        "top5_acc",
        "top10_acc",
    ]

    for col, header in enumerate(detail_headers, start=2):
        write_cell(ws, detail_header_row, col, header, detail_header_style)

    ordered_df = run_df.copy()
    ordered_df = ordered_df.reset_index(drop=True)
    row_idx = detail_data_start
    for seq, (_, row) in enumerate(ordered_df.iterrows(), start=1):
        full_metrics = [
            row.get("best_valid_full_ndcg@10"),
            row.get("best_valid_full_hr@10"),
            row.get("best_valid_full_ndcg@5"),
            row.get("best_valid_full_hr@5"),
            row.get("best_valid_full_mrr"),
            row.get("best_valid_task_accuracy"),
            row.get("best_valid_task_macro_f1"),
            row.get("best_valid_task_top5_accuracy"),
            row.get("best_valid_task_top10_accuracy"),
            row.get("best_test_at_best_valid_full_ndcg@10"),
            row.get("best_test_at_best_valid_full_hr@10"),
            row.get("best_test_at_best_valid_full_ndcg@5"),
            row.get("best_test_at_best_valid_full_hr@5"),
            row.get("best_test_at_best_valid_full_mrr"),
            row.get("best_test_at_best_valid_task_accuracy"),
            row.get("best_test_at_best_valid_task_macro_f1"),
            row.get("best_test_at_best_valid_task_top5_accuracy"),
            row.get("best_test_at_best_valid_task_top10_accuracy"),
        ]
        sampled_metrics = [
            row.get("best_valid_sampled_ndcg@10"),
            row.get("best_valid_sampled_hr@10"),
            row.get("best_valid_sampled_ndcg@5"),
            row.get("best_valid_sampled_hr@5"),
            row.get("best_valid_sampled_mrr"),
            row.get("best_valid_task_accuracy"),
            row.get("best_valid_task_macro_f1"),
            row.get("best_valid_task_top5_accuracy"),
            row.get("best_valid_task_top10_accuracy"),
            row.get("best_test_at_best_valid_sampled_ndcg@10"),
            row.get("best_test_at_best_valid_sampled_hr@10"),
            row.get("best_test_at_best_valid_sampled_ndcg@5"),
            row.get("best_test_at_best_valid_sampled_hr@5"),
            row.get("best_test_at_best_valid_sampled_mrr"),
            row.get("best_test_at_best_valid_task_accuracy"),
            row.get("best_test_at_best_valid_task_macro_f1"),
            row.get("best_test_at_best_valid_task_top5_accuracy"),
            row.get("best_test_at_best_valid_task_top10_accuracy"),
        ]
        common = [
            "BPI 2012",
            row["variant"],
            row["run_name"],
            row["maxlen"],
            row["dropout_rate"],
            "NDCG@10",
        ]
        full_row = [seq * 2 - 1] + common[:2] + [common[2], row["seed"], common[3], common[4], row["time_encoding"], row["평가 방식"] if "평가 방식" in row else "full ranking", common[5]] + full_metrics
        sampled_row = [seq * 2] + common[:2] + [common[2], row["seed"], common[3], common[4], row["time_encoding"], "negative sampling(100)", common[5]] + sampled_metrics
        for values in (full_row, sampled_row):
            for col, value in enumerate(values, start=2):
                write_cell(ws, row_idx, col, value, text_style)
            row_idx += 1

    # Mean / std section
    meanstd_start = row_idx + 2
    write_cell(ws, meanstd_start, 2, "▶ run별 성능 평균, 표준편차", section_title_style)

    meanstd_group_row = meanstd_start + 1
    meanstd_metric_row = meanstd_start + 2
    meanstd_header_row = meanstd_start + 3
    meanstd_data_start = meanstd_start + 4

    metric_start_col = 9
    group_defs = [
        ("VALID set 성능", 9),
        ("TEST set 성능", 9),
    ]

    col_ptr = metric_start_col
    for group_name, metric_count in group_defs:
        span = metric_count * 2
        ws.merge_cells(start_row=meanstd_group_row, start_column=col_ptr, end_row=meanstd_group_row, end_column=col_ptr + span - 1)
        write_cell(ws, meanstd_group_row, col_ptr, group_name, meanstd_group_style)
        col_ptr += span

    metric_label_groups = [
        ["NDCG@10", "Hit@10", "NDCG@5", "Hit@5", "MRR", "accuracy", "macro_f1", "top5_acc", "top10_acc"],
        ["NDCG@10", "Hit@10", "NDCG@5", "Hit@5", "MRR", "accuracy", "macro_f1", "top5_acc", "top10_acc"],
    ]
    col_ptr = metric_start_col
    for labels in metric_label_groups:
        for label in labels:
            ws.merge_cells(start_row=meanstd_metric_row, start_column=col_ptr, end_row=meanstd_metric_row, end_column=col_ptr + 1)
            write_cell(ws, meanstd_metric_row, col_ptr, label, meanstd_subheader_style)
            write_cell(ws, meanstd_header_row, col_ptr, "mean", meanstd_header_style)
            write_cell(ws, meanstd_header_row, col_ptr + 1, "std", meanstd_header_style)
            col_ptr += 2

    id_headers = ["SEQ", "Dataset", "variant", "run_name", "maxlen", "dropout_rate", "평가 방식", "selection_metric"]
    for col, header in enumerate(id_headers, start=2):
        write_cell(ws, meanstd_header_row, col, header, meanstd_header_style)

    summary_row_start = meanstd_data_start
    for seq, variant in enumerate(SUMMARY_VARIANTS, start=1):
        base_values = [
            seq * 2 - 1,
            "BPI 2012",
            variant,
            variant,
            ordered_df.loc[ordered_df["variant"] == variant, "maxlen"].iloc[0],
            ordered_df.loc[ordered_df["variant"] == variant, "dropout_rate"].iloc[0],
            "full ranking",
            "NDCG@10",
        ]
        for col, value in enumerate(base_values, start=2):
            write_cell(ws, summary_row_start, col, value, text_style)

        col_ptr = metric_start_col
        full_summary_metrics = [
            "best_valid_full_ndcg@10", "best_valid_full_hr@10", "best_valid_full_ndcg@5", "best_valid_full_hr@5", "best_valid_full_mrr",
            "best_valid_task_accuracy", "best_valid_task_macro_f1", "best_valid_task_top5_accuracy", "best_valid_task_top10_accuracy",
            "best_test_at_best_valid_full_ndcg@10", "best_test_at_best_valid_full_hr@10", "best_test_at_best_valid_full_ndcg@5", "best_test_at_best_valid_full_hr@5", "best_test_at_best_valid_full_mrr",
            "best_test_at_best_valid_task_accuracy", "best_test_at_best_valid_task_macro_f1", "best_test_at_best_valid_task_top5_accuracy", "best_test_at_best_valid_task_top10_accuracy",
        ]
        for metric in full_summary_metrics:
            write_cell(ws, summary_row_start, col_ptr, summary_df.loc[variant, (metric, "mean")], text_style)
            write_cell(ws, summary_row_start, col_ptr + 1, summary_df.loc[variant, (metric, "std")], text_style)
            col_ptr += 2

        sampled_row_start = summary_row_start + 1
        sampled_base_values = [
            seq * 2,
            "BPI 2012",
            variant,
            variant,
            ordered_df.loc[ordered_df["variant"] == variant, "maxlen"].iloc[0],
            ordered_df.loc[ordered_df["variant"] == variant, "dropout_rate"].iloc[0],
            "negative sampling(100)",
            "NDCG@10",
        ]
        for col, value in enumerate(sampled_base_values, start=2):
            write_cell(ws, sampled_row_start, col, value, text_style)
        col_ptr = metric_start_col
        sampled_summary_metrics = [
            "best_valid_sampled_ndcg@10", "best_valid_sampled_hr@10", "best_valid_sampled_ndcg@5", "best_valid_sampled_hr@5", "best_valid_sampled_mrr",
            "best_valid_task_accuracy", "best_valid_task_macro_f1", "best_valid_task_top5_accuracy", "best_valid_task_top10_accuracy",
            "best_test_at_best_valid_sampled_ndcg@10", "best_test_at_best_valid_sampled_hr@10", "best_test_at_best_valid_sampled_ndcg@5", "best_test_at_best_valid_sampled_hr@5", "best_test_at_best_valid_sampled_mrr",
            "best_test_at_best_valid_task_accuracy", "best_test_at_best_valid_task_macro_f1", "best_test_at_best_valid_task_top5_accuracy", "best_test_at_best_valid_task_top10_accuracy",
        ]
        for metric in sampled_summary_metrics:
            write_cell(ws, sampled_row_start, col_ptr, summary_df.loc[variant, (metric, "mean")], text_style)
            write_cell(ws, sampled_row_start, col_ptr + 1, summary_df.loc[variant, (metric, "std")], text_style)
            col_ptr += 2
        summary_row_start += 2

    try:
        wb.save(WORKBOOK_PATH)
        print(f"[ok] updated workbook: {WORKBOOK_PATH}")
    except PermissionError:
        fallback_path = FALLBACK_WORKBOOK_PATH
        try:
            wb.save(fallback_path)
            print(f"[warn] original workbook was locked, wrote fallback file: {fallback_path}")
        except PermissionError:
            stamped = ROOT / "docs" / f"임시_stage2_sinusoidal_added_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
            wb.save(stamped)
            print(f"[warn] workbook files were locked, wrote timestamped fallback file: {stamped}")


if __name__ == "__main__":
    main()
