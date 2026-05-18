from __future__ import annotations

import io
import json
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
NOTEBOOKS = ROOT / "notebooks"
DOCS = ROOT / "docs"
TEMPLATE_PATH = DOCS / "모델성능.xlsx"
DETAIL_SOURCE_PATH = DOCS / "timeaware_results_02_08_260517.xlsx"
OUTPUT_PATH = DOCS / "모델성능_02_08_정리_260517.xlsx"

NOTEBOOK_FILES = [
    "sasrec_timeaware_bpi2012_colab_train_02_260513.ipynb",
    "sasrec_timeaware_bpi2012_colab_train_03_260513.ipynb",
    "sasrec_timeaware_bpi2012_colab_train_04_260514.ipynb",
    "sasrec_timeaware_bpi2012_colab_train_05_260514.ipynb",
    "sasrec_timeaware_bpi2012_colab_train_06_260514.ipynb",
    "sasrec_timeaware_bpi2012_colab_train_07_260514.ipynb",
    "sasrec_timeaware_bpi2012_colab_train_08_260515.ipynb",
]

SHEET_META = {
    "sasrec_timeaware_bpi2012_colab_train_02_260513.ipynb": {
        "title": "02_anchor_additive",
        "model_desc": "anchor_ml20 baseline + time-aware additive",
        "note": [
            "- baseline: anchor_ml20",
            "- time-aware: 각 이벤트 입력벡터에 시간 정보를 더하는 방식",
            "- time source: 직전 이벤트와의 시간 차이(delta_prev_seconds)",
            "- compare: bucket 2종류(b8, b9)",
        ],
    },
    "sasrec_timeaware_bpi2012_colab_train_03_260513.ipynb": {
        "title": "03_refine_bucket",
        "model_desc": "refine_ml50_do035 baseline + time-aware additive",
        "note": [
            "- baseline: refine_ml50_do035",
            "- time-aware: 각 이벤트 입력벡터에 시간 정보를 더하는 방식",
            "- time source: 직전 이벤트와의 시간 차이(delta_prev_seconds)",
            "- compare: bucket 2종류(b8, b9)",
        ],
    },
    "sasrec_timeaware_bpi2012_colab_train_04_260514.ipynb": {
        "title": "04_refine_conti",
        "model_desc": "refine_ml50_do035 baseline + time-aware additive",
        "note": [
            "- baseline: refine_ml50_do035",
            "- time-aware: 각 이벤트 입력벡터에 시간 정보를 더하는 방식",
            "- time source: 직전 이벤트와의 시간 차이(delta_prev_seconds)",
            "- compare: continuous time encoding",
        ],
    },
    "sasrec_timeaware_bpi2012_colab_train_05_260514.ipynb": {
        "title": "05_dstart_add",
        "model_desc": "refine_ml50_do035 baseline + time-aware additive",
        "note": [
            "- baseline: refine_ml50_do035",
            "- time-aware: 각 이벤트 입력벡터에 시간 정보를 더하는 방식",
            "- time source: case 시작 이후 누적 시간(delta_start_seconds)",
            "- compare: 9-bucket vs continuous",
        ],
    },
    "sasrec_timeaware_bpi2012_colab_train_06_260514.ipynb": {
        "title": "06_attnbias_refine",
        "model_desc": "refine_ml50_do035 baseline + time-aware attention bias",
        "note": [
            "- baseline: refine_ml50_do035",
            "- time-aware: attention score에 시간 bias를 추가하는 방식",
            "- time source: case 시작 이후 누적 시간(delta_start_seconds)",
            "- compare: 9-bucket attention bias",
        ],
    },
    "sasrec_timeaware_bpi2012_colab_train_07_260514.ipynb": {
        "title": "07_attnbias_anchor",
        "model_desc": "anchor_ml20 baseline + time-aware attention bias",
        "note": [
            "- baseline: anchor_ml20",
            "- time-aware: attention score에 시간 bias를 추가하는 방식",
            "- time source: case 시작 이후 누적 시간(delta_start_seconds)",
            "- compare: 9-bucket attention bias",
        ],
    },
    "sasrec_timeaware_bpi2012_colab_train_08_260515.ipynb": {
        "title": "08_attnbias_b10",
        "model_desc": "refine_ml50_do035 baseline + time-aware attention bias",
        "note": [
            "- baseline: refine_ml50_do035",
            "- time-aware: attention score에 시간 bias를 추가하는 방식",
            "- time source: case 시작 이후 누적 시간(delta_start_seconds)",
            "- compare: 기존 b9와 짧은 구간을 더 나눈 b10 비교",
        ],
    },
}

NUMERIC_COLS = [
    "best_valid_full_ndcg@10",
    "best_valid_full_hr@10",
    "best_valid_full_ndcg@5",
    "best_valid_full_hr@5",
    "best_valid_full_mrr",
    "best_test_full_ndcg@10",
    "best_test_full_hr@10",
    "best_test_full_ndcg@5",
    "best_test_full_hr@5",
    "best_test_full_mrr",
    "best_valid_sampled_ndcg@10",
    "best_valid_sampled_hr@10",
    "best_valid_sampled_ndcg@5",
    "best_valid_sampled_hr@5",
    "best_valid_sampled_mrr",
    "best_test_sampled_ndcg@10",
    "best_test_sampled_hr@10",
    "best_test_sampled_ndcg@5",
    "best_test_sampled_hr@5",
    "best_test_sampled_mrr",
]


def parse_detail_table_from_output(output: dict) -> pd.DataFrame:
    text = "".join(output.get("data", {}).get("text/plain", []))
    df = pd.read_fwf(io.StringIO(text))

    if "Unnamed: 0" in df.columns:
        df = df.drop(columns=["Unnamed: 0"])

    if "bucket_variant" in df.columns and "time_variant" not in df.columns:
        df = df.rename(columns={"bucket_variant": "time_variant"})
    if "time_bucket_boundaries" in df.columns and "time_bucket_boundaries_parsed" not in df.columns:
        df = df.rename(columns={"time_bucket_boundaries": "time_bucket_boundaries_parsed"})

    unnamed_cols = [c for c in df.columns if str(c).startswith("Unnamed:")]
    if "time_bucket_boundaries_parsed" in df.columns and unnamed_cols:
        for col in unnamed_cols:
            if df[col].notna().any():
                df["time_bucket_boundaries_parsed"] = (
                    df[col].fillna("").astype(str).str.strip()
                    + " "
                    + df["time_bucket_boundaries_parsed"].fillna("").astype(str).str.strip()
                ).str.strip()
        df = df.drop(columns=unnamed_cols)

    for col in NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def extract_detail_frames(notebook_path: Path) -> list[pd.DataFrame]:
    nb = json.loads(notebook_path.read_text(encoding="utf-8"))
    frames: list[pd.DataFrame] = []

    for cell in nb["cells"]:
        if cell.get("cell_type") != "code":
            continue

        source = "".join(cell.get("source", []))
        selection_basis = None
        if "df_ndcg10[[" in source:
            selection_basis = "NDCG@10"
        elif "df_ndcg5[[" in source:
            selection_basis = "NDCG@5"

        if selection_basis is None:
            continue

        table_output = next(
            (o for o in cell.get("outputs", []) if "data" in o and "text/plain" in o.get("data", {})),
            None,
        )
        if table_output is None:
            continue

        df = parse_detail_table_from_output(table_output)
        df["notebook"] = notebook_path.name
        df["selection_basis"] = selection_basis
        frames.append(df)

    return frames


def build_detail_df() -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for name in NOTEBOOK_FILES:
        path = NOTEBOOKS / name
        if path.exists():
            frames.extend(extract_detail_frames(path))
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def build_summary_df(detail_df: pd.DataFrame) -> pd.DataFrame:
    summary = (
        detail_df.groupby(["notebook", "selection_basis", "time_variant"], dropna=False)[NUMERIC_COLS]
        .agg(["mean", "std"])
        .reset_index()
    )
    summary.columns = [
        f"{metric}_{agg}" if agg else metric
        for metric, agg in [
            (col if isinstance(col, str) else col[0], "" if isinstance(col, str) else col[1])
            for col in summary.columns
        ]
    ]
    return summary


def expand_detail_rows(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in df.iterrows():
        for eval_mode, prefix in [("full ranking", "full"), ("negative sampling(100)", "sampled")]:
            rows.append(
                {
                    "Dataset": "BPI 2012",
                    "run_name": row.get("run_name"),
                    "maxlen": row.get("maxlen"),
                    "dropout_rate": row.get("dropout_rate"),
                    "평가 방식": eval_mode,
                    "best epoch 기준": row.get("selection_basis"),
                    "NDCG@10_valid": row.get(f"best_valid_{prefix}_ndcg@10"),
                    "Hit@10_valid": row.get(f"best_valid_{prefix}_hr@10"),
                    "NDCG@5_valid": row.get(f"best_valid_{prefix}_ndcg@5"),
                    "Hit@5_valid": row.get(f"best_valid_{prefix}_hr@5"),
                    "MRR_valid": row.get(f"best_valid_{prefix}_mrr"),
                    "NDCG@10_test": row.get(f"best_test_{prefix}_ndcg@10"),
                    "Hit@10_test": row.get(f"best_test_{prefix}_hr@10"),
                    "NDCG@5_test": row.get(f"best_test_{prefix}_ndcg@5"),
                    "Hit@5_test": row.get(f"best_test_{prefix}_hr@5"),
                    "MRR_test": row.get(f"best_test_{prefix}_mrr"),
                }
            )
    expanded = pd.DataFrame(rows)
    return expanded


def expand_summary_rows(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in df.iterrows():
        for eval_mode, prefix in [("full ranking", "full"), ("negative sampling(100)", "sampled")]:
            rows.append(
                {
                    "Dataset": "BPI 2012",
                    "run_name": row.get("time_variant"),
                    "maxlen": None,
                    "dropout_rate": None,
                    "평가 방식": eval_mode,
                    "best epoch 기준": row.get("selection_basis"),
                    "NDCG@10_valid_mean": row.get(f"best_valid_{prefix}_ndcg@10_mean"),
                    "NDCG@10_valid_std": row.get(f"best_valid_{prefix}_ndcg@10_std"),
                    "Hit@10_valid_mean": row.get(f"best_valid_{prefix}_hr@10_mean"),
                    "Hit@10_valid_std": row.get(f"best_valid_{prefix}_hr@10_std"),
                    "NDCG@5_valid_mean": row.get(f"best_valid_{prefix}_ndcg@5_mean"),
                    "NDCG@5_valid_std": row.get(f"best_valid_{prefix}_ndcg@5_std"),
                    "Hit@5_valid_mean": row.get(f"best_valid_{prefix}_hr@5_mean"),
                    "Hit@5_valid_std": row.get(f"best_valid_{prefix}_hr@5_std"),
                    "MRR_valid_mean": row.get(f"best_valid_{prefix}_mrr_mean"),
                    "MRR_valid_std": row.get(f"best_valid_{prefix}_mrr_std"),
                    "NDCG@10_test_mean": row.get(f"best_test_{prefix}_ndcg@10_mean"),
                    "NDCG@10_test_std": row.get(f"best_test_{prefix}_ndcg@10_std"),
                    "Hit@10_test_mean": row.get(f"best_test_{prefix}_hr@10_mean"),
                    "Hit@10_test_std": row.get(f"best_test_{prefix}_hr@10_std"),
                    "NDCG@5_test_mean": row.get(f"best_test_{prefix}_ndcg@5_mean"),
                    "NDCG@5_test_std": row.get(f"best_test_{prefix}_ndcg@5_std"),
                    "Hit@5_test_mean": row.get(f"best_test_{prefix}_hr@5_mean"),
                    "Hit@5_test_std": row.get(f"best_test_{prefix}_hr@5_std"),
                    "MRR_test_mean": row.get(f"best_test_{prefix}_mrr_mean"),
                    "MRR_test_std": row.get(f"best_test_{prefix}_mrr_std"),
                }
            )
    return pd.DataFrame(rows)


def copy_row_style(ws, source_row: int, target_row: int, max_col: int = 28) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)


def set_num(cell, value) -> None:
    cell.value = None if pd.isna(value) else float(value)


def infer_config_from_run_name(run_name: str) -> tuple[int | None, float | None]:
    if not isinstance(run_name, str):
        return None, None

    maxlen = None
    dropout = None

    if "ml100" in run_name:
        maxlen = 100
    elif "ml75" in run_name:
        maxlen = 75
    elif "ml50" in run_name:
        maxlen = 50
    elif "ml20" in run_name:
        maxlen = 20
    elif "pd" in run_name:
        maxlen = 200

    if "do035" in run_name:
        dropout = 0.35
    elif "do030" in run_name:
        dropout = 0.30
    elif "do025" in run_name:
        dropout = 0.25
    elif "anchor_ml20" in run_name or "attnbias_dstart_ml20" in run_name or "timeaware_anchor_ml20" in run_name:
        dropout = 0.20

    return maxlen, dropout


def fill_sheet(ws, notebook_name: str, detail_df: pd.DataFrame, summary_df: pd.DataFrame) -> None:
    meta = SHEET_META[notebook_name]
    sheet_detail = detail_df[detail_df["notebook"] == notebook_name].copy()
    sheet_summary = summary_df[summary_df["notebook"] == notebook_name].copy()

    expanded_detail = expand_detail_rows(sheet_detail)
    expanded_summary = expand_summary_rows(sheet_summary)

    base_detail_capacity = 24  # rows 15~38
    detail_extra = max(0, len(expanded_detail) - base_detail_capacity)
    if detail_extra:
        ws.insert_rows(39, amount=detail_extra)
        for row_idx in range(39, 39 + detail_extra):
            copy_row_style(ws, 38, row_idx)

    summary_start = 44 + detail_extra
    base_summary_capacity = 8
    summary_extra = max(0, len(expanded_summary) - base_summary_capacity)
    if summary_extra:
        ws.insert_rows(summary_start + base_summary_capacity, amount=summary_extra)
        for row_idx in range(summary_start + base_summary_capacity, summary_start + base_summary_capacity + summary_extra):
            copy_row_style(ws, summary_start + base_summary_capacity - 1, row_idx)

    # Template merged ranges do not follow inserted rows reliably, so rebuild
    # the summary header merges at their effective row.
    for merged in ["I41:R41", "S41:AB41"]:
        if merged in {str(rng) for rng in ws.merged_cells.ranges}:
            ws.unmerge_cells(merged)
    summary_header_row = 41 + detail_extra
    ws.merge_cells(start_row=summary_header_row, start_column=9, end_row=summary_header_row, end_column=18)
    ws.merge_cells(start_row=summary_header_row, start_column=19, end_row=summary_header_row, end_column=28)

    ws["B2"] = "SUMMARY"
    ws["B3"] = f"- 모델 구조: {meta['model_desc']}"
    for row_idx in range(4, 13):
        ws.cell(row_idx, 2).value = None
    for offset, line in enumerate(meta["note"], start=5):
        ws.cell(offset, 2).value = line

    variant_names = ", ".join(dict.fromkeys(str(v) for v in sheet_summary["time_variant"].dropna().tolist()))
    ws["B10"] = f"- variants: {variant_names}"

    baseline_row = sheet_summary[
        (sheet_summary["selection_basis"] == "NDCG@10") & (sheet_summary["time_variant"] == "baseline")
    ]
    non_baseline = sheet_summary[
        (sheet_summary["selection_basis"] == "NDCG@10") & (sheet_summary["time_variant"] != "baseline")
    ]
    best_row = non_baseline.sort_values("best_test_full_ndcg@10_mean", ascending=False).head(1)

    if not baseline_row.empty and not best_row.empty:
        b = baseline_row.iloc[0]
        t = best_row.iloc[0]
        diff = float(t["best_test_full_ndcg@10_mean"]) - float(b["best_test_full_ndcg@10_mean"])
        direction = "상회" if diff > 0 else "하회"
        ws["B11"] = (
            f"- main metric(full ranking, NDCG@10) 기준 best time-aware는 {t['time_variant']} "
            f"(test mean {t['best_test_full_ndcg@10_mean']:.4f})"
        )
        ws["B12"] = (
            f"  baseline(test mean {b['best_test_full_ndcg@10_mean']:.4f}) 대비 "
            f"{abs(diff):.4f} {direction}"
        )
        ws["G5"] = "- 아래 성능은 full ranking, NDCG@10 기준 핵심 비교"
        ws["G6"] = f"- baseline : {b['best_test_full_ndcg@10_mean']:.4f} / {b['best_test_full_ndcg@10_std']:.4f}"
        ws["G7"] = f"- best time-aware({t['time_variant']}) : {t['best_test_full_ndcg@10_mean']:.4f} / {t['best_test_full_ndcg@10_std']:.4f}"
        ws["G8"] = f"- valid(test) NDCG@10 : {t['best_valid_full_ndcg@10_mean']:.4f} / {t['best_test_full_ndcg@10_mean']:.4f}"
    else:
        ws["B11"] = "- baseline 또는 time-aware summary row를 찾지 못했습니다."

    # Detail rows
    detail_start = 15
    for idx, (_, row) in enumerate(expanded_detail.iterrows(), start=detail_start):
        seq = idx - detail_start + 1
        inferred_maxlen, inferred_dropout = infer_config_from_run_name(row["run_name"])
        maxlen = row["maxlen"] if pd.notna(row["maxlen"]) else inferred_maxlen
        dropout = row["dropout_rate"] if pd.notna(row["dropout_rate"]) else inferred_dropout
        ws.cell(idx, 2).value = seq
        ws.cell(idx, 3).value = row["Dataset"]
        ws.cell(idx, 4).value = row["run_name"]
        ws.cell(idx, 5).value = maxlen
        ws.cell(idx, 6).value = dropout
        ws.cell(idx, 7).value = row["평가 방식"]
        ws.cell(idx, 8).value = row["best epoch 기준"]
        set_num(ws.cell(idx, 9), row["NDCG@10_valid"])
        set_num(ws.cell(idx, 10), row["Hit@10_valid"])
        set_num(ws.cell(idx, 11), row["NDCG@5_valid"])
        set_num(ws.cell(idx, 12), row["Hit@5_valid"])
        set_num(ws.cell(idx, 13), row["MRR_valid"])
        set_num(ws.cell(idx, 14), row["NDCG@10_test"])
        set_num(ws.cell(idx, 15), row["Hit@10_test"])
        set_num(ws.cell(idx, 16), row["NDCG@5_test"])
        set_num(ws.cell(idx, 17), row["Hit@5_test"])
        set_num(ws.cell(idx, 18), row["MRR_test"])

    # Summary rows
    for idx, (_, row) in enumerate(expanded_summary.iterrows(), start=summary_start):
        seq = idx - summary_start + 1
        ws.cell(idx, 2).value = seq
        ws.cell(idx, 3).value = row["Dataset"]
        ws.cell(idx, 4).value = row["run_name"]
        maxlen = None
        dropout = None
        if row["run_name"] == "baseline" and not baseline_row.empty:
            base_runs = sheet_detail[sheet_detail["time_variant"] == "baseline"]
            if not base_runs.empty:
                maxlen = base_runs.iloc[0].get("maxlen")
                dropout = base_runs.iloc[0].get("dropout_rate")
        ws.cell(idx, 5).value = maxlen
        ws.cell(idx, 6).value = dropout
        ws.cell(idx, 7).value = row["평가 방식"]
        ws.cell(idx, 8).value = row["best epoch 기준"]
        set_num(ws.cell(idx, 9), row["NDCG@10_valid_mean"])
        set_num(ws.cell(idx, 10), row["NDCG@10_valid_std"])
        set_num(ws.cell(idx, 11), row["Hit@10_valid_mean"])
        set_num(ws.cell(idx, 12), row["Hit@10_valid_std"])
        set_num(ws.cell(idx, 13), row["NDCG@5_valid_mean"])
        set_num(ws.cell(idx, 14), row["NDCG@5_valid_std"])
        set_num(ws.cell(idx, 15), row["Hit@5_valid_mean"])
        set_num(ws.cell(idx, 16), row["Hit@5_valid_std"])
        set_num(ws.cell(idx, 17), row["MRR_valid_mean"])
        set_num(ws.cell(idx, 18), row["MRR_valid_std"])
        set_num(ws.cell(idx, 19), row["NDCG@10_test_mean"])
        set_num(ws.cell(idx, 20), row["NDCG@10_test_std"])
        set_num(ws.cell(idx, 21), row["Hit@10_test_mean"])
        set_num(ws.cell(idx, 22), row["Hit@10_test_std"])
        set_num(ws.cell(idx, 23), row["NDCG@5_test_mean"])
        set_num(ws.cell(idx, 24), row["NDCG@5_test_std"])
        set_num(ws.cell(idx, 25), row["Hit@5_test_mean"])
        set_num(ws.cell(idx, 26), row["Hit@5_test_std"])
        set_num(ws.cell(idx, 27), row["MRR_test_mean"])
        set_num(ws.cell(idx, 28), row["MRR_test_std"])

    max_row = summary_start + len(expanded_summary) + 2
    for row in ws.iter_rows(min_row=15, max_row=max_row, min_col=9, max_col=28):
        for cell in row:
            cell.number_format = "0.0000"


def main() -> None:
    detail_df = build_detail_df()
    summary_df = build_summary_df(detail_df)

    wb = load_workbook(TEMPLATE_PATH)
    template = wb[wb.sheetnames[0]]
    copies = {}
    for notebook_name in NOTEBOOK_FILES[1:]:
        ws = wb.copy_worksheet(template)
        ws.title = SHEET_META[notebook_name]["title"]
        copies[notebook_name] = ws

    first_name = NOTEBOOK_FILES[0]
    template.title = SHEET_META[first_name]["title"]
    copies[first_name] = template

    for notebook_name in NOTEBOOK_FILES:
        fill_sheet(copies[notebook_name], notebook_name, detail_df, summary_df)

    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
