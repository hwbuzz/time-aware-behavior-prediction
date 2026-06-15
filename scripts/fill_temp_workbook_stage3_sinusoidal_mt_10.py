from __future__ import annotations

from copy import copy
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path

import nbformat
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "docs" / "임시_stage2_sinusoidal_added_20260614_132150.xlsx"
NOTEBOOK_PATH = ROOT / "notebooks" / "sasrec_stage3_bpi2012_colab_train_10_260614.ipynb"
SOURCE_SHEET_NAME = "2-3.Sinusoidal"
TARGET_SHEET_NAME = "3-7.Sinusoidal_MT"

DETAIL_METRICS = [
    "best_valid_full_ndcg@10",
    "best_valid_full_hr@10",
    "best_valid_full_ndcg@5",
    "best_valid_full_hr@5",
    "best_valid_full_mrr",
    "best_test_at_best_valid_full_ndcg@10",
    "best_test_at_best_valid_full_hr@10",
    "best_test_at_best_valid_full_ndcg@5",
    "best_test_at_best_valid_full_hr@5",
    "best_test_at_best_valid_full_mrr",
    "best_valid_sampled_ndcg@10",
    "best_valid_sampled_hr@10",
    "best_valid_sampled_ndcg@5",
    "best_valid_sampled_hr@5",
    "best_valid_sampled_mrr",
    "best_test_at_best_valid_sampled_ndcg@10",
    "best_test_at_best_valid_sampled_hr@10",
    "best_test_at_best_valid_sampled_ndcg@5",
    "best_test_at_best_valid_sampled_hr@5",
    "best_test_at_best_valid_sampled_mrr",
    "best_valid_task_accuracy",
    "best_valid_task_macro_f1",
    "best_valid_task_top5_accuracy",
    "best_valid_task_top10_accuracy",
    "best_valid_task_time_mae",
    "best_valid_task_time_rmse",
    "best_valid_task_time_median_ae",
    "best_test_at_best_valid_task_accuracy",
    "best_test_at_best_valid_task_macro_f1",
    "best_test_at_best_valid_task_top5_accuracy",
    "best_test_at_best_valid_task_top10_accuracy",
    "best_test_at_best_valid_task_time_mae",
    "best_test_at_best_valid_task_time_rmse",
    "best_test_at_best_valid_task_time_median_ae",
]

SUMMARY_VARIANTS = [
    "refine_baseline",
    "refine_attnbias_single_task",
    "refine_sinusoidal_single_task",
    "refine_multi_task_w1.0",
    "refine_attnbias_multi_task_w1.0",
    "refine_sinusoidal_multi_task_w1.0",
    "refine_sinusoidal_multi_task_w0.1",
]

METRIC_LABELS = [
    "NDCG@10",
    "Hit@10",
    "NDCG@5",
    "Hit@5",
    "MRR",
    "Accuracy",
    "MacroF1",
    "Top5Acc",
    "Top10Acc",
    "MAE",
    "RMSE",
    "MedianAE",
]


class SimpleTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._current_row: list[str] = []
        self._in_cell = False
        self._buf = ""

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self._current_row = []
        elif tag in ("th", "td"):
            self._in_cell = True
            self._buf = ""

    def handle_endtag(self, tag):
        if tag in ("th", "td") and self._in_cell:
            self._current_row.append(self._buf.strip())
            self._in_cell = False
        elif tag == "tr" and self._current_row:
            self.rows.append(self._current_row)

    def handle_data(self, data):
        if self._in_cell:
            self._buf += data


def convert_value(value):
    if value in ("", "None", "nan", "NaN", None):
        return None
    if isinstance(value, str):
        low = value.lower()
        if low == "true":
            return True
        if low == "false":
            return False
        try:
            if "." not in value and "e" not in low:
                return int(value)
            return float(value)
        except ValueError:
            return value
    return value


def parse_notebook_run_table() -> pd.DataFrame:
    nb = nbformat.read(NOTEBOOK_PATH, as_version=4)
    html = nb.cells[35]["outputs"][0]["data"]["text/html"]
    parser = SimpleTableParser()
    parser.feed(html)

    header = parser.rows[0][1:]
    records = []
    for row in parser.rows[1:]:
        values = row[1:]
        if len(values) != len(header):
            continue
        records.append({k: convert_value(v) for k, v in zip(header, values)})
    df = pd.DataFrame(records)
    return df.sort_values(["variant", "seed", "run_name"]).reset_index(drop=True)


def copy_style(src, dst) -> None:
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def write_cell(ws, row: int, col: int, value, style_cell=None):
    cell = ws.cell(row=row, column=col, value=value)
    if style_cell is not None:
        copy_style(style_cell, cell)
    return cell


def col_letter(idx: int) -> str:
    result = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def full_metric_keys(prefix: str) -> list[str]:
    return [
        f"{prefix}_full_ndcg@10",
        f"{prefix}_full_hr@10",
        f"{prefix}_full_ndcg@5",
        f"{prefix}_full_hr@5",
        f"{prefix}_full_mrr",
        f"{prefix}_task_accuracy",
        f"{prefix}_task_macro_f1",
        f"{prefix}_task_top5_accuracy",
        f"{prefix}_task_top10_accuracy",
        f"{prefix}_task_time_mae",
        f"{prefix}_task_time_rmse",
        f"{prefix}_task_time_median_ae",
    ]


def sampled_metric_keys(prefix: str) -> list[str]:
    return [
        f"{prefix}_sampled_ndcg@10",
        f"{prefix}_sampled_hr@10",
        f"{prefix}_sampled_ndcg@5",
        f"{prefix}_sampled_hr@5",
        f"{prefix}_sampled_mrr",
        f"{prefix}_task_accuracy",
        f"{prefix}_task_macro_f1",
        f"{prefix}_task_top5_accuracy",
        f"{prefix}_task_top10_accuracy",
        f"{prefix}_task_time_mae",
        f"{prefix}_task_time_rmse",
        f"{prefix}_task_time_median_ae",
    ]


def set_widths(ws):
    widths = {
        "B": 6, "C": 12, "D": 28, "E": 34, "F": 8, "G": 8, "H": 10, "I": 12, "J": 22, "K": 12, "L": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for idx in range(13, 90):
        ws.column_dimensions[col_letter(idx)].width = 12


def main() -> None:
    df = parse_notebook_run_table()
    summary_df = df.groupby("variant")[DETAIL_METRICS].agg(["mean", "std"]).reindex(SUMMARY_VARIANTS)

    wb = load_workbook(WORKBOOK_PATH)
    src = wb[SOURCE_SHEET_NAME]
    if TARGET_SHEET_NAME in wb.sheetnames:
        del wb[TARGET_SHEET_NAME]
    ws = wb.create_sheet(TARGET_SHEET_NAME)
    ws.freeze_panes = "A18"
    set_widths(ws)

    title_style = src["B2"]
    text_style = src["B4"]
    label_style = src["C6"]
    small_header_style = src["I8"]
    group_header_style = src["L59"]
    detail_header_style = src["B60"]
    section_title_style = src["B54"]
    meanstd_group_style = src["H100"]
    meanstd_subheader_style = src["H101"]
    meanstd_header_style = src["B102"]

    write_cell(ws, 2, 2, "SUMMARY", title_style)
    write_cell(ws, 4, 2, " refine backbone Stage 3 sinusoidal multi-task results", text_style)
    write_cell(ws, 4, 10, "main metric(full ranking, NDCG@10) 기준", text_style)
    write_cell(ws, 5, 10, "  single-task sinusoidal(test mean 0.8644) > sinusoidal multi-task w0.1(0.8257) > sinusoidal multi-task w1.0(0.7705)", text_style)
    write_cell(ws, 6, 3, "sinusoidal multitask", label_style)
    write_cell(ws, 6, 4, "delta_start + sinusoidal input embedding + next activity/next time joint learning", text_style)
    write_cell(ws, 6, 10, "  plain multitask(0.7112) 대비 sinusoidal multitask는 activity mean은 개선", text_style)
    write_cell(ws, 7, 3, "time_loss_weight", label_style)
    write_cell(ws, 7, 4, "w1.0, w0.1 비교", text_style)
    write_cell(ws, 8, 3, "time target", label_style)
    write_cell(ws, 8, 4, "delta_next_seconds, log1p, huber", text_style)
    write_cell(ws, 10, 3, "비교 대상", label_style)
    write_cell(ws, 10, 4, "refine baseline, attnbias single-task, sinusoidal single-task, plain/attnbias/sinusoidal multitask", text_style)

    summary_headers = ["variant", "split", "metric", "mean", "std"]
    for offset, header in enumerate(summary_headers, start=10):
        write_cell(ws, 8, offset, header, small_header_style)

    summary_rows = [
        ("refine_baseline", "test", "full_ndcg@10"),
        ("refine_attnbias_single_task", "test", "full_ndcg@10"),
        ("refine_sinusoidal_single_task", "test", "full_ndcg@10"),
        ("refine_multi_task_w1.0", "test", "full_ndcg@10"),
        ("refine_attnbias_multi_task_w1.0", "test", "full_ndcg@10"),
        ("refine_sinusoidal_multi_task_w1.0", "test", "full_ndcg@10"),
        ("refine_sinusoidal_multi_task_w0.1", "test", "full_ndcg@10"),
        ("refine_multi_task_w1.0", "test", "task_time_mae"),
        ("refine_attnbias_multi_task_w1.0", "test", "task_time_mae"),
        ("refine_sinusoidal_multi_task_w1.0", "test", "task_time_mae"),
        ("refine_sinusoidal_multi_task_w0.1", "test", "task_time_mae"),
    ]
    metric_map = {
        "full_ndcg@10": "best_test_at_best_valid_full_ndcg@10",
        "task_time_mae": "best_test_at_best_valid_task_time_mae",
    }
    row_ptr = 9
    for variant, split, label in summary_rows:
        metric = metric_map[label]
        write_cell(ws, row_ptr, 10, variant, text_style)
        write_cell(ws, row_ptr, 11, split, text_style)
        write_cell(ws, row_ptr, 12, label, text_style)
        write_cell(ws, row_ptr, 13, summary_df.loc[variant, (metric, "mean")], text_style)
        write_cell(ws, row_ptr, 14, summary_df.loc[variant, (metric, "std")], text_style)
        row_ptr += 1

    detail_start = row_ptr + 2
    group_row = detail_start
    header_row = detail_start + 1
    data_start = detail_start + 2

    detail_groups = [
        ("VALID set 성능", 12),
        ("TEST set 성능", 12),
    ]

    start_col = 13
    col_ptr = start_col
    for name, count in detail_groups:
        ws.merge_cells(start_row=group_row, start_column=col_ptr, end_row=group_row, end_column=col_ptr + count - 1)
        write_cell(ws, group_row, col_ptr, name, group_header_style)
        col_ptr += count

    headers = [
        "SEQ", "Dataset", "variant", "run_name", "seed", "maxlen", "dropout_rate", "time_encoding",
        "time_delta_column", "time_loss_weight", "평가 방식", "best epoch 기준",
        "NDCG@10", "Hit@10", "NDCG@5", "Hit@5", "MRR", "Accuracy", "MacroF1", "Top5Acc", "Top10Acc", "MAE", "RMSE", "MedianAE",
        "NDCG@10", "Hit@10", "NDCG@5", "Hit@5", "MRR", "Accuracy", "MacroF1", "Top5Acc", "Top10Acc", "MAE", "RMSE", "MedianAE",
    ]
    for col, header in enumerate(headers, start=2):
        write_cell(ws, header_row, col, header, detail_header_style)

    for seq, (_, row) in enumerate(df.iterrows(), start=1):
        full_metrics = [row.get(key) for key in full_metric_keys("best_valid")] + [row.get(key) for key in full_metric_keys("best_test_at_best_valid")]
        sampled_metrics = [row.get(key) for key in sampled_metric_keys("best_valid")] + [row.get(key) for key in sampled_metric_keys("best_test_at_best_valid")]
        base = [ "BPI 2012", row["variant"], row["run_name"], row["seed"], row["maxlen"], row["dropout_rate"], row["time_encoding"], row["time_delta_column"], row["time_loss_weight"] ]
        full_row = [seq * 2 - 1] + base + ["full ranking", "NDCG@10"] + full_metrics
        sampled_row = [seq * 2] + base + ["negative sampling(100)", "NDCG@10"] + sampled_metrics
        for values in (full_row, sampled_row):
            for col, value in enumerate(values, start=2):
                write_cell(ws, data_start, col, value, text_style)
            data_start += 1

    mean_start = data_start + 2
    write_cell(ws, mean_start, 2, "▶ run별 성능 평균, 표준편차", section_title_style)

    mean_group_row = mean_start + 1
    mean_metric_row = mean_start + 2
    mean_header_row = mean_start + 3
    mean_data_start = mean_start + 4

    mean_groups = [
        ("VALID set 성능", 12),
        ("TEST set 성능", 12),
    ]
    metric_labels = [METRIC_LABELS, METRIC_LABELS]
    metric_col_start = 10
    col_ptr = metric_col_start
    for name, count in mean_groups:
        span = count * 2
        ws.merge_cells(start_row=mean_group_row, start_column=col_ptr, end_row=mean_group_row, end_column=col_ptr + span - 1)
        write_cell(ws, mean_group_row, col_ptr, name, meanstd_group_style)
        col_ptr += span

    col_ptr = metric_col_start
    for labels in metric_labels:
        for label in labels:
            ws.merge_cells(start_row=mean_metric_row, start_column=col_ptr, end_row=mean_metric_row, end_column=col_ptr + 1)
            write_cell(ws, mean_metric_row, col_ptr, label, meanstd_subheader_style)
            write_cell(ws, mean_header_row, col_ptr, "mean", meanstd_header_style)
            write_cell(ws, mean_header_row, col_ptr + 1, "std", meanstd_header_style)
            col_ptr += 2

    for col, header in enumerate(["SEQ", "Dataset", "variant", "run_name", "maxlen", "dropout_rate", "평가 방식", "selection_metric"], start=2):
        write_cell(ws, mean_header_row, col, header, meanstd_header_style)

    row_ptr = mean_data_start
    for seq, variant in enumerate(SUMMARY_VARIANTS, start=1):
        sample = df[df["variant"] == variant].iloc[0]
        base_values = [seq * 2 - 1, "BPI 2012", variant, variant, sample["maxlen"], sample["dropout_rate"], "full ranking", "NDCG@10"]
        full_metrics = full_metric_keys("best_valid") + full_metric_keys("best_test_at_best_valid")
        col_ptr = metric_col_start
        target_row = row_ptr
        for col, value in enumerate(base_values, start=2):
            write_cell(ws, target_row, col, value, text_style)
        for metric in full_metrics:
            write_cell(ws, target_row, col_ptr, summary_df.loc[variant, (metric, "mean")], text_style)
            write_cell(ws, target_row, col_ptr + 1, summary_df.loc[variant, (metric, "std")], text_style)
            col_ptr += 2
        sampled_base = [seq * 2, "BPI 2012", variant, variant, sample["maxlen"], sample["dropout_rate"], "negative sampling(100)", "NDCG@10"]
        for col, value in enumerate(sampled_base, start=2):
            write_cell(ws, target_row + 1, col, value, text_style)
        sampled_metrics = sampled_metric_keys("best_valid") + sampled_metric_keys("best_test_at_best_valid")
        col_ptr = metric_col_start
        for metric in sampled_metrics:
            write_cell(ws, target_row + 1, col_ptr, summary_df.loc[variant, (metric, "mean")], text_style)
            write_cell(ws, target_row + 1, col_ptr + 1, summary_df.loc[variant, (metric, "std")], text_style)
            col_ptr += 2
        row_ptr += 2

    try:
        wb.save(WORKBOOK_PATH)
        print(f"[ok] updated workbook: {WORKBOOK_PATH}")
    except PermissionError:
        stamped = ROOT / "docs" / f"임시_stage3_sinusoidal_mt_added_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        wb.save(stamped)
        print(f"[warn] workbook was locked, wrote timestamped fallback file: {stamped}")


if __name__ == "__main__":
    main()
