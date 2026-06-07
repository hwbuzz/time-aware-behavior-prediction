from __future__ import annotations

import json
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from shutil import copy2
from typing import Iterable

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter


WORKBOOK_PATH = Path("docs/캡스톤_모델성능_v0.5_260604.xlsx")

STAGE3_NOTEBOOKS = {
    "02": Path("notebooks/sasrec_stage3_bpi2012_colab_train_02_260601.ipynb"),
    "03": Path("notebooks/sasrec_stage3_bpi2012_colab_train_03_260601.ipynb"),
    "04": Path("notebooks/sasrec_stage3_bpi2012_colab_train_04_260601.ipynb"),
    "05": Path("notebooks/sasrec_stage3_bpi2012_colab_train_05_260602.ipynb"),
    "06": Path("notebooks/sasrec_stage3_bpi2012_colab_train_06_260602.ipynb"),
    "07": Path("notebooks/sasrec_stage3_bpi2012_time_naive_baselines_07_260602.ipynb"),
}

NEW_SHEETS = [
    "3-0.Stage3",
    "3-1.Baseline_mt",
    "3-2.AttnBias_mt",
    "3-3.AttnBias_w01",
    "3-4.Anchor_w01",
    "3-5.Refine_w01",
    "3-6.Time_naive",
]

SUMMARY_METRICS = [
    "best_valid_full_ndcg@10",
    "best_valid_full_hr@10",
    "best_valid_full_ndcg@5",
    "best_valid_full_hr@5",
    "best_valid_full_mrr",
    "best_valid_sampled_ndcg@10",
    "best_valid_sampled_hr@10",
    "best_valid_sampled_ndcg@5",
    "best_valid_sampled_hr@5",
    "best_valid_sampled_mrr",
    "best_valid_task_accuracy",
    "best_valid_task_macro_f1",
    "best_valid_task_top5_accuracy",
    "best_valid_task_top10_accuracy",
    "best_valid_task_time_mae",
    "best_valid_task_time_rmse",
    "best_valid_task_time_median_ae",
    "best_test_at_best_valid_full_ndcg@10",
    "best_test_at_best_valid_full_hr@10",
    "best_test_at_best_valid_full_ndcg@5",
    "best_test_at_best_valid_full_hr@5",
    "best_test_at_best_valid_full_mrr",
    "best_test_at_best_valid_sampled_ndcg@10",
    "best_test_at_best_valid_sampled_hr@10",
    "best_test_at_best_valid_sampled_ndcg@5",
    "best_test_at_best_valid_sampled_hr@5",
    "best_test_at_best_valid_sampled_mrr",
    "best_test_at_best_valid_task_accuracy",
    "best_test_at_best_valid_task_macro_f1",
    "best_test_at_best_valid_task_top5_accuracy",
    "best_test_at_best_valid_task_top10_accuracy",
    "best_test_at_best_valid_task_time_mae",
    "best_test_at_best_valid_task_time_rmse",
    "best_test_at_best_valid_task_time_median_ae",
]

RAW_METRIC_ORDER = SUMMARY_METRICS + [
    "last_valid_full_ndcg@10",
    "last_valid_full_hr@10",
    "last_valid_full_ndcg@5",
    "last_valid_full_hr@5",
    "last_valid_full_mrr",
    "last_valid_sampled_ndcg@10",
    "last_valid_sampled_hr@10",
    "last_valid_sampled_ndcg@5",
    "last_valid_sampled_hr@5",
    "last_valid_sampled_mrr",
    "last_valid_task_accuracy",
    "last_valid_task_macro_f1",
    "last_valid_task_top5_accuracy",
    "last_valid_task_top10_accuracy",
    "last_valid_task_time_mae",
    "last_valid_task_time_rmse",
    "last_valid_task_time_median_ae",
    "last_test_full_ndcg@10",
    "last_test_full_hr@10",
    "last_test_full_ndcg@5",
    "last_test_full_hr@5",
    "last_test_full_mrr",
    "last_test_sampled_ndcg@10",
    "last_test_sampled_hr@10",
    "last_test_sampled_ndcg@5",
    "last_test_sampled_hr@5",
    "last_test_sampled_mrr",
    "last_test_task_accuracy",
    "last_test_task_macro_f1",
    "last_test_task_top5_accuracy",
    "last_test_task_top10_accuracy",
    "last_test_task_time_mae",
    "last_test_task_time_rmse",
    "last_test_task_time_median_ae",
]


class SimpleHTMLTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._in_cell = False
        self._text = ""
        self._colspan = 1

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attr_map = dict(attrs)
        if tag == "tr":
            self._row = []
        elif tag in {"th", "td"}:
            self._in_cell = True
            self._text = ""
            self._colspan = int(attr_map.get("colspan", "1") or "1")

    def handle_endtag(self, tag: str) -> None:
        if tag in {"th", "td"} and self._in_cell:
            text = " ".join(self._text.split())
            if self._row is not None:
                self._row.extend([text] * self._colspan)
            self._in_cell = False
        elif tag == "tr" and self._row is not None:
            self.rows.append(self._row)
            self._row = None

    def handle_data(self, data: str) -> None:
        if self._in_cell:
            self._text += data


@dataclass
class NotebookTables:
    detail: pd.DataFrame | None = None
    summary: pd.DataFrame | None = None
    naive_readable: pd.DataFrame | None = None
    model_summary: pd.DataFrame | None = None


def parse_html_rows(html_parts: str | list[str]) -> list[list[str]]:
    html = "".join(html_parts) if isinstance(html_parts, list) else html_parts
    parser = SimpleHTMLTableParser()
    parser.feed(html)
    return parser.rows


def rows_to_df(rows: list[list[str]]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame()

    if len(rows) >= 3 and len(rows[1]) > 1 and rows[1][1] in {"mean", "std"}:
        index_name = rows[2][0] or rows[0][0] or "index"
        columns = [index_name]
        for metric, stat in zip(rows[0][1:], rows[1][1:]):
            columns.append(f"{metric}|{stat}")
        data = [r[: len(columns)] for r in rows[3:]]
        return pd.DataFrame(data, columns=columns)

    header = rows[0]
    data = [r[: len(header)] for r in rows[1:]]
    return pd.DataFrame(data, columns=header)


def load_notebook_tables(nb_path: Path) -> NotebookTables:
    nb = json.loads(nb_path.read_text(encoding="utf-8"))
    result = NotebookTables()

    for cell in nb["cells"]:
        source = "".join(cell.get("source", []))
        outputs = cell.get("outputs", [])
        html_outputs = []
        for output in outputs:
            data = output.get("data", {})
            if "text/html" in data:
                html_outputs.append(data["text/html"])
        if not html_outputs:
            continue

        if "df_compare[existing_display_cols]" in source or "df_compare[display_cols]" in source:
            result.detail = rows_to_df(parse_html_rows(html_outputs[0]))
        elif "summary_compare =" in source:
            result.summary = rows_to_df(parse_html_rows(html_outputs[0]))
        elif "readable" in source and "baseline_results" in source:
            result.naive_readable = rows_to_df(parse_html_rows(html_outputs[0]))
        elif "def rebuild_df" in source:
            result.model_summary = rows_to_df(parse_html_rows(html_outputs[0]))

    return result


def maybe_numeric(value: str):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    if text in {"None", "nan", "NaN"}:
        return None
    try:
        if "." in text or "e" in text.lower():
            return float(text)
        return int(text)
    except ValueError:
        return text


def clean_detail_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    if "" in out.columns:
        out = out.drop(columns=[""])
    for col in out.columns:
        out[col] = out[col].map(maybe_numeric)
    return out


def clean_summary_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    index_col = out.columns[0]
    out = out.rename(columns={index_col: "variant"})
    for col in out.columns[1:]:
        out[col] = out[col].map(maybe_numeric)
    filtered_cols = ["variant"]
    for metric in SUMMARY_METRICS:
        for stat in ("mean", "std"):
            key = f"{metric}|{stat}"
            if key in out.columns:
                filtered_cols.append(key)
    return out[filtered_cols]


def clean_naive_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    if "" in out.columns:
        out = out.drop(columns=[""])
    for col in out.columns:
        out[col] = out[col].map(maybe_numeric)
    return out


def ordered_metric_columns(available_cols: Iterable[str], metric_order: list[str]) -> list[str]:
    available = list(available_cols)
    ordered = [c for c in metric_order if c in available]
    remaining = [c for c in available if c not in ordered]
    return ordered + remaining


RANKING_BLOCKS = [
    ("full ranking", "full"),
    ("negative sampling(100)", "sampled"),
]

COMMON_VALID_TASK_METRICS = [
    "best_valid_task_accuracy",
    "best_valid_task_macro_f1",
    "best_valid_task_top5_accuracy",
    "best_valid_task_top10_accuracy",
    "best_valid_task_time_mae",
    "best_valid_task_time_rmse",
    "best_valid_task_time_median_ae",
]

COMMON_TEST_TASK_METRICS = [
    "best_test_at_best_valid_task_accuracy",
    "best_test_at_best_valid_task_macro_f1",
    "best_test_at_best_valid_task_top5_accuracy",
    "best_test_at_best_valid_task_top10_accuracy",
    "best_test_at_best_valid_task_time_mae",
    "best_test_at_best_valid_task_time_rmse",
    "best_test_at_best_valid_task_time_median_ae",
]

DISPLAY_METRIC_LABELS = {
    "best_valid_task_accuracy": "VALID Accuracy",
    "best_valid_task_macro_f1": "VALID MacroF1",
    "best_valid_task_top5_accuracy": "VALID Top5Acc",
    "best_valid_task_top10_accuracy": "VALID Top10Acc",
    "best_valid_task_time_mae": "VALID Time MAE",
    "best_valid_task_time_rmse": "VALID Time RMSE",
    "best_valid_task_time_median_ae": "VALID Time MedianAE",
    "best_test_at_best_valid_task_accuracy": "TEST Accuracy",
    "best_test_at_best_valid_task_macro_f1": "TEST MacroF1",
    "best_test_at_best_valid_task_top5_accuracy": "TEST Top5Acc",
    "best_test_at_best_valid_task_top10_accuracy": "TEST Top10Acc",
    "best_test_at_best_valid_task_time_mae": "TEST Time MAE",
    "best_test_at_best_valid_task_time_rmse": "TEST Time RMSE",
    "best_test_at_best_valid_task_time_median_ae": "TEST Time MedianAE",
}

SUMMARY_METRIC_LAYOUT = [
    ("NDCG@10", "VALID NDCG@10 mean", "VALID NDCG@10 std", "TEST NDCG@10 mean", "TEST NDCG@10 std"),
    ("Hit@10", "VALID Hit@10 mean", "VALID Hit@10 std", "TEST Hit@10 mean", "TEST Hit@10 std"),
    ("NDCG@5", "VALID NDCG@5 mean", "VALID NDCG@5 std", "TEST NDCG@5 mean", "TEST NDCG@5 std"),
    ("Hit@5", "VALID Hit@5 mean", "VALID Hit@5 std", "TEST Hit@5 mean", "TEST Hit@5 std"),
    ("MRR", "VALID MRR mean", "VALID MRR std", "TEST MRR mean", "TEST MRR std"),
    ("Accuracy", "VALID Accuracy mean", "VALID Accuracy std", "TEST Accuracy mean", "TEST Accuracy std"),
    ("MacroF1", "VALID MacroF1 mean", "VALID MacroF1 std", "TEST MacroF1 mean", "TEST MacroF1 std"),
    ("Top5Acc", "VALID Top5Acc mean", "VALID Top5Acc std", "TEST Top5Acc mean", "TEST Top5Acc std"),
    ("Top10Acc", "VALID Top10Acc mean", "VALID Top10Acc std", "TEST Top10Acc mean", "TEST Top10Acc std"),
    ("Time MAE", "VALID Time MAE mean", "VALID Time MAE std", "TEST Time MAE mean", "TEST Time MAE std"),
    ("Time RMSE", "VALID Time RMSE mean", "VALID Time RMSE std", "TEST Time RMSE mean", "TEST Time RMSE std"),
    ("Time MedianAE", "VALID Time MedianAE mean", "VALID Time MedianAE std", "TEST Time MedianAE mean", "TEST Time MedianAE std"),
]


def build_split_raw_table(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    rows = []
    for _, row in df.iterrows():
        for eval_label, prefix in RANKING_BLOCKS:
            out = {
                "Dataset": "BPI 2012",
                "run_name": row.get("run_name"),
                "seed": row.get("seed"),
                "variant": row.get("variant"),
                "maxlen": row.get("maxlen"),
                "dropout_rate": row.get("dropout_rate"),
                "selection_metric": row.get("selection_metric"),
                "time_loss_weight": row.get("time_loss_weight"),
                "평가 방식": eval_label,
                "best epoch 기준": row.get("selection_metric"),
                "VALID NDCG@10": row.get(f"best_valid_{prefix}_ndcg@10"),
                "VALID Hit@10": row.get(f"best_valid_{prefix}_hr@10"),
                "VALID NDCG@5": row.get(f"best_valid_{prefix}_ndcg@5"),
                "VALID Hit@5": row.get(f"best_valid_{prefix}_hr@5"),
                "VALID MRR": row.get(f"best_valid_{prefix}_mrr"),
            }
            for metric in COMMON_VALID_TASK_METRICS:
                out[DISPLAY_METRIC_LABELS[metric]] = row.get(metric)
            out["TEST NDCG@10"] = row.get(f"best_test_at_best_valid_{prefix}_ndcg@10")
            out["TEST Hit@10"] = row.get(f"best_test_at_best_valid_{prefix}_hr@10")
            out["TEST NDCG@5"] = row.get(f"best_test_at_best_valid_{prefix}_ndcg@5")
            out["TEST Hit@5"] = row.get(f"best_test_at_best_valid_{prefix}_hr@5")
            out["TEST MRR"] = row.get(f"best_test_at_best_valid_{prefix}_mrr")
            for metric in COMMON_TEST_TASK_METRICS:
                out[DISPLAY_METRIC_LABELS[metric]] = row.get(metric)
            rows.append(out)

    out_df = pd.DataFrame(rows)
    out_df.insert(0, "SEQ", range(1, len(out_df) + 1))
    return out_df


def build_split_summary_table(summary_df: pd.DataFrame, detail_df: pd.DataFrame) -> pd.DataFrame:
    if summary_df.empty:
        return pd.DataFrame()

    meta_cols = ["variant", "maxlen", "dropout_rate", "time_loss_weight"]
    meta_df = detail_df[[c for c in meta_cols if c in detail_df.columns]].drop_duplicates(subset=["variant"])
    meta_lookup = {
        row["variant"]: row.to_dict()
        for _, row in meta_df.iterrows()
    }

    rows = []
    for _, row in summary_df.iterrows():
        variant = row["variant"]
        meta = meta_lookup.get(variant, {})
        for eval_label, prefix in RANKING_BLOCKS:
            out = {
                "Dataset": "BPI 2012",
                "variant": variant,
                "maxlen": meta.get("maxlen"),
                "dropout_rate": meta.get("dropout_rate"),
                "time_loss_weight": meta.get("time_loss_weight"),
                "평가 방식": eval_label,
                "best epoch 기준": "NDCG@10",
                "VALID NDCG@10 mean": row.get(f"best_valid_{prefix}_ndcg@10|mean"),
                "VALID NDCG@10 std": row.get(f"best_valid_{prefix}_ndcg@10|std"),
                "VALID Hit@10 mean": row.get(f"best_valid_{prefix}_hr@10|mean"),
                "VALID Hit@10 std": row.get(f"best_valid_{prefix}_hr@10|std"),
                "VALID NDCG@5 mean": row.get(f"best_valid_{prefix}_ndcg@5|mean"),
                "VALID NDCG@5 std": row.get(f"best_valid_{prefix}_ndcg@5|std"),
                "VALID Hit@5 mean": row.get(f"best_valid_{prefix}_hr@5|mean"),
                "VALID Hit@5 std": row.get(f"best_valid_{prefix}_hr@5|std"),
                "VALID MRR mean": row.get(f"best_valid_{prefix}_mrr|mean"),
                "VALID MRR std": row.get(f"best_valid_{prefix}_mrr|std"),
            }
            for metric in COMMON_VALID_TASK_METRICS:
                label = DISPLAY_METRIC_LABELS[metric]
                out[f"{label} mean"] = row.get(f"{metric}|mean")
                out[f"{label} std"] = row.get(f"{metric}|std")
            out["TEST NDCG@10 mean"] = row.get(f"best_test_at_best_valid_{prefix}_ndcg@10|mean")
            out["TEST NDCG@10 std"] = row.get(f"best_test_at_best_valid_{prefix}_ndcg@10|std")
            out["TEST Hit@10 mean"] = row.get(f"best_test_at_best_valid_{prefix}_hr@10|mean")
            out["TEST Hit@10 std"] = row.get(f"best_test_at_best_valid_{prefix}_hr@10|std")
            out["TEST NDCG@5 mean"] = row.get(f"best_test_at_best_valid_{prefix}_ndcg@5|mean")
            out["TEST NDCG@5 std"] = row.get(f"best_test_at_best_valid_{prefix}_ndcg@5|std")
            out["TEST Hit@5 mean"] = row.get(f"best_test_at_best_valid_{prefix}_hr@5|mean")
            out["TEST Hit@5 std"] = row.get(f"best_test_at_best_valid_{prefix}_hr@5|std")
            out["TEST MRR mean"] = row.get(f"best_test_at_best_valid_{prefix}_mrr|mean")
            out["TEST MRR std"] = row.get(f"best_test_at_best_valid_{prefix}_mrr|std")
            for metric in COMMON_TEST_TASK_METRICS:
                label = DISPLAY_METRIC_LABELS[metric]
                out[f"{label} mean"] = row.get(f"{metric}|mean")
                out[f"{label} std"] = row.get(f"{metric}|std")
            rows.append(out)

    out_df = pd.DataFrame(rows)
    out_df.insert(0, "SEQ", range(1, len(out_df) + 1))
    return out_df


def get_metric(summary_df: pd.DataFrame, variant: str, metric: str, stat: str = "mean"):
    if summary_df.empty:
        return None
    key = f"{metric}|{stat}"
    rows = summary_df[summary_df["variant"] == variant]
    if rows.empty or key not in rows.columns:
        return None
    return rows.iloc[0][key]


def fnum(value, digits: int = 4) -> str:
    if value is None or value == "":
        return "-"
    if isinstance(value, str):
        return value
    return f"{float(value):.{digits}f}"


def fsec_to_hours(value) -> str:
    if value is None or value == "":
        return "-"
    return f"{float(value) / 3600:.2f}h"


def clone_cell_style(src, dst) -> None:
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def apply_template_styles(ws, template_ws) -> None:
    clone_cell_style(template_ws["B2"], ws["B2"])
    clone_cell_style(template_ws["B4"], ws["B4"])
    clone_cell_style(template_ws["C6"], ws["C6"])
    clone_cell_style(template_ws["D6"], ws["D6"])
    clone_cell_style(template_ws["B12"], ws["B12"])
    clone_cell_style(template_ws["B13"], ws["B13"])
    for col in range(2, 60):
        clone_cell_style(template_ws.cell(16, min(col, template_ws.max_column)), ws.cell(16, col))
        clone_cell_style(template_ws.cell(57, min(col, template_ws.max_column)), ws.cell(57, col))
    clone_cell_style(template_ws["I55"], ws["B55"])
    clone_cell_style(template_ws["S55"], ws["T55"])


def set_default_layout(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B17"
    ws.column_dimensions["A"].width = 3
    for col in range(2, 60):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18


def write_summary_block(ws, title: str, rows: list[tuple[str, str]], bullets: list[str], max_col: int = 28) -> int:
    end_col_letter = get_column_letter(max_col)
    ws["B2"] = "SUMMARY"
    ws.merge_cells(f"B4:{end_col_letter}4")
    ws["B4"] = title

    current_row = 6
    for key, value in rows:
        ws[f"C{current_row}"] = key
        ws[f"D{current_row}"] = value
        current_row += 1

    current_row += 1
    for bullet in bullets:
        ws.merge_cells(f"B{current_row}:{end_col_letter}{current_row}")
        ws[f"B{current_row}"] = f"▶ {bullet}"
        current_row += 1
    return current_row + 1


def write_dataframe(ws, start_row: int, start_col: int, df: pd.DataFrame, title: str | None = None) -> int:
    row = start_row
    if title:
        ws.cell(row, start_col, title)
        ws.cell(row, start_col).font = Font(bold=True)
        row += 1

    for idx, col in enumerate(df.columns, start_col):
        cell = ws.cell(row, idx, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=copy(ws["B16"].border.left),
            right=copy(ws["B16"].border.right),
            top=copy(ws["B16"].border.top),
            bottom=copy(ws["B16"].border.bottom),
        )
    row += 1

    for _, record in df.iterrows():
        for idx, col in enumerate(df.columns, start_col):
            val = record[col]
            cell = ws.cell(row, idx, val)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row += 1
    return row


def write_stage2_like_summary_table(ws, start_row: int, start_col: int, df: pd.DataFrame, title: str) -> int:
    title_end_col = start_col + 7 + (len(SUMMARY_METRIC_LAYOUT) * 4) - 1
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=title_end_col)
    ws.cell(start_row, start_col, title)
    ws.cell(start_row, start_col).font = Font(bold=True)

    group_row = start_row + 1
    metric_row = start_row + 2
    stat_row = start_row + 3
    data_start = start_row + 4

    meta_headers = ["SEQ", "Dataset", "variant", "maxlen", "dropout_rate", "time_loss_weight", "평가 방식", "best epoch 기준"]
    for idx, header in enumerate(meta_headers, start_col):
        cell = ws.cell(stat_row, idx, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=copy(ws["B16"].border.left),
            right=copy(ws["B16"].border.right),
            top=copy(ws["B16"].border.top),
            bottom=copy(ws["B16"].border.bottom),
        )

    valid_start = start_col + len(meta_headers)
    valid_end = valid_start + (len(SUMMARY_METRIC_LAYOUT) * 2) - 1
    test_start = valid_end + 1
    test_end = test_start + (len(SUMMARY_METRIC_LAYOUT) * 2) - 1

    ws.merge_cells(start_row=group_row, start_column=valid_start, end_row=group_row, end_column=valid_end)
    ws.merge_cells(start_row=group_row, start_column=test_start, end_row=group_row, end_column=test_end)
    ws.cell(group_row, valid_start, "VALID set 성능")
    ws.cell(group_row, test_start, "TEST set 성능")

    for c in range(valid_start, test_end + 1):
        ws.cell(group_row, c).font = Font(bold=True)
        ws.cell(group_row, c).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(group_row, c).alignment = Alignment(horizontal="center", vertical="center")

    current_col = valid_start
    for metric_label, _, _, _, _ in SUMMARY_METRIC_LAYOUT:
        ws.merge_cells(start_row=metric_row, start_column=current_col, end_row=metric_row, end_column=current_col + 1)
        ws.cell(metric_row, current_col, metric_label)
        ws.cell(stat_row, current_col, "mean")
        ws.cell(stat_row, current_col + 1, "std")
        current_col += 2

    for c in range(valid_start, valid_end + 1):
        ws.cell(metric_row, c).font = Font(bold=True)
        ws.cell(metric_row, c).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(metric_row, c).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(stat_row, c).font = Font(bold=True)
        ws.cell(stat_row, c).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(stat_row, c).alignment = Alignment(horizontal="center", vertical="center")

    current_col = test_start
    for metric_label, _, _, _, _ in SUMMARY_METRIC_LAYOUT:
        ws.merge_cells(start_row=metric_row, start_column=current_col, end_row=metric_row, end_column=current_col + 1)
        ws.cell(metric_row, current_col, metric_label)
        ws.cell(stat_row, current_col, "mean")
        ws.cell(stat_row, current_col + 1, "std")
        current_col += 2

    for c in range(test_start, test_end + 1):
        ws.cell(metric_row, c).font = Font(bold=True)
        ws.cell(metric_row, c).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(metric_row, c).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(stat_row, c).font = Font(bold=True)
        ws.cell(stat_row, c).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(stat_row, c).alignment = Alignment(horizontal="center", vertical="center")

    for row_offset, (_, row) in enumerate(df.iterrows()):
        excel_row = data_start + row_offset
        values = [
            row.get("SEQ"),
            row.get("Dataset"),
            row.get("variant"),
            row.get("maxlen"),
            row.get("dropout_rate"),
            row.get("time_loss_weight"),
            row.get("평가 방식"),
            row.get("best epoch 기준"),
        ]
        for idx, value in enumerate(values, start_col):
            ws.cell(excel_row, idx, value)
            ws.cell(excel_row, idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        current_col = valid_start
        for _, v_mean, v_std, t_mean, t_std in SUMMARY_METRIC_LAYOUT:
            ws.cell(excel_row, current_col, row.get(v_mean))
            ws.cell(excel_row, current_col + 1, row.get(v_std))
            current_col += 2
        current_col = test_start
        for _, v_mean, v_std, t_mean, t_std in SUMMARY_METRIC_LAYOUT:
            ws.cell(excel_row, current_col, row.get(t_mean))
            ws.cell(excel_row, current_col + 1, row.get(t_std))
            current_col += 2

        for c in range(start_col, test_end + 1):
            ws.cell(excel_row, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    return data_start + len(df)


def build_overview_df(parsed: dict[str, NotebookTables]) -> pd.DataFrame:
    rows = []
    s02 = clean_summary_df(parsed["02"].summary)
    s03 = clean_summary_df(parsed["03"].summary)
    s04 = clean_summary_df(parsed["04"].summary)
    s05 = clean_summary_df(parsed["05"].summary)
    s06 = clean_summary_df(parsed["06"].summary)

    key_variants = [
        ("02", "anchor_single_task", "single-task baseline"),
        ("02", "anchor_multi_task", "plain multitask"),
        ("03", "anchor_attnbias_multi_task", "attnbias multitask"),
        ("05", "anchor_multi_task_w0.1", "plain multitask w0.1"),
        ("06", "refine_multi_task_w0.1", "refine multitask w0.1"),
    ]
    for exp_no, variant, note in key_variants:
        sdf = {"02": s02, "03": s03, "04": s04, "05": s05, "06": s06}[exp_no]
        rows.append(
            {
                "실험": exp_no,
                "variant": variant,
                "설명": note,
                "test_full_ndcg@10_mean": get_metric(sdf, variant, "best_test_at_best_valid_full_ndcg@10", "mean"),
                "test_full_ndcg@10_std": get_metric(sdf, variant, "best_test_at_best_valid_full_ndcg@10", "std"),
                "test_time_mae_mean": get_metric(sdf, variant, "best_test_at_best_valid_task_time_mae", "mean"),
                "test_time_mae_std": get_metric(sdf, variant, "best_test_at_best_valid_task_time_mae", "std"),
            }
        )
    return pd.DataFrame(rows)


def create_stage3_overview(ws, template_ws, parsed: dict[str, NotebookTables]) -> None:
    apply_template_styles(ws, template_ws)
    set_default_layout(ws)

    s02 = clean_summary_df(parsed["02"].summary)
    s03 = clean_summary_df(parsed["03"].summary)
    s05 = clean_summary_df(parsed["05"].summary)
    s06 = clean_summary_df(parsed["06"].summary)

    # Rebuild top block in the style of 2-0.TimeAware.
    ws["B2"] = "Stage 3 Multitask 실험"
    ws["C3"] = "SASRec backbone 위에 next activity classification head와 next time regression head를 함께 두고 multitask learning으로 학습함"

    info_rows = {
        5: ("B", "0. 예측 target / 구조"),
        6: ("C", "- next activity"),
        7: ("D", "다음 이벤트의 activity 예측"),
        8: ("C", "- next time (delta_next_seconds)"),
        9: ("D", "현재 이벤트와 다음 이벤트 사이의 시간 차이(초 단위) 예측"),
        10: ("D", "shared sequence representation 위에 activity head와 time head를 함께 학습"),
        12: ("B", "1. Loss / 학습 설정"),
        13: ("C", "activity loss + time loss를 함께 사용"),
        14: ("C", "time loss는 Huber loss를 사용"),
        15: ("C", "time target은 log1p(delta_next_seconds) 변환 후 학습"),
        16: ("C", "time_loss_weight를 조정하여 activity-time trade-off 확인"),
        18: ("B", "2. Main metric"),
        19: ("C", "- activity main metric"),
        20: ("D", "full ranking, NDCG@10"),
        21: ("C", "- time main metric"),
        22: ("D", "MAE"),
        23: ("C", "- 보조 metric"),
        24: ("D", "RMSE, Median AE"),
        26: ("B", "3. 비교 축"),
        27: ("C", "- backbone"),
        28: ("D", "anchor_ml20, refine_ml50_do035"),
        29: ("C", "- multitask 구조"),
        30: ("D", "plain multitask, attention-bias multitask"),
        31: ("C", "- loss weight"),
        32: ("D", "time_loss_weight = 1.0, 0.1"),
        34: ("B", "4. 해석 포인트"),
        35: ("C", "single-task에서 multi-task로 갈 때 next activity 성능이 얼마나 유지되는지 확인"),
        36: ("C", "time_loss_weight 조절에 따라 activity와 time prediction 사이 trade-off가 어떻게 바뀌는지 확인"),
        37: ("C", "next-time 해석을 위해 global mean/median, activity mean, prefix length mean naive baseline과 비교"),
    }
    for row_idx, (col, value) in info_rows.items():
        ws[f"{col}{row_idx}"] = value

    # Match 2-0 style by copying a few key styles.
    style_src = wb_style_source = None
    # use 2-0.TimeAware sheet when available for visual consistency
    if "2-0.TimeAware" in ws.parent.sheetnames:
        wb_style_source = ws.parent["2-0.TimeAware"]
        for ref, target in [
            ("B2", "B2"),
            ("C3", "C3"),
            ("B5", "B5"),
            ("B11", "B12"),
            ("C6", "C6"),
            ("D7", "D7"),
        ]:
            clone_cell_style(wb_style_source[ref], ws[target])
        for row_idx in [5, 12, 18, 26, 34]:
            clone_cell_style(wb_style_source["B5"], ws[f"B{row_idx}"])
        for row_idx in [6, 8, 13, 14, 15, 16, 19, 21, 23, 27, 29, 31, 35, 36, 37]:
            clone_cell_style(wb_style_source["C6"], ws[f"C{row_idx}"])
        for row_idx in [7, 9, 10, 20, 22, 24, 28, 30, 32]:
            clone_cell_style(wb_style_source["D7"], ws[f"D{row_idx}"])

    # concise bottom summary
    ws["B40"] = "요약"
    ws["B41"] = (
        "single-task에서 multi-task로 가면 next activity 성능이 전반적으로 낮아졌고, "
        "time_loss_weight를 0.1로 낮추면 activity 성능은 회복되지만 time MAE와 trade-off가 발생함."
    )
    ws["B42"] = (
        "Stage 3 best multitask(activity 기준)는 anchor_multi_task_w0.1 "
        f"(test full NDCG@10={fnum(get_metric(s05, 'anchor_multi_task_w0.1', 'best_test_at_best_valid_full_ndcg@10', 'mean'))})이며, "
        "best time MAE는 anchor_attnbias_multi_task "
        f"(MAE={fnum(get_metric(s03, 'anchor_attnbias_multi_task', 'best_test_at_best_valid_task_time_mae', 'mean'), 1)}초)임."
    )
    ws["B43"] = (
        "refine에서도 w0.1 조정 효과가 재현되었지만, overall main metric 기준으로는 "
        f"anchor_multi_task_w0.1 ({fnum(get_metric(s05, 'anchor_multi_task_w0.1', 'best_test_at_best_valid_full_ndcg@10', 'mean'))})"
        f" > refine_multi_task_w0.1 ({fnum(get_metric(s06, 'refine_multi_task_w0.1', 'best_test_at_best_valid_full_ndcg@10', 'mean'))})."
    )
    if wb_style_source is not None:
        clone_cell_style(wb_style_source["B5"], ws["B40"])
        clone_cell_style(template_ws["B12"], ws["B41"])
        clone_cell_style(template_ws["B12"], ws["B42"])
        clone_cell_style(template_ws["B12"], ws["B43"])

    overview_df = build_overview_df(parsed)
    next_row = write_dataframe(ws, 46, 2, overview_df, "핵심 variant 비교")

    naive = clean_naive_df(parsed["07"].naive_readable)
    if not naive.empty:
        test_naive = naive[naive["split"] == "test"].copy()
        test_naive = test_naive[
            [
                "split",
                "baseline",
                "mae",
                "rmse",
                "median_ae",
                "mae_hours",
                "mae_minutes",
                "rmse_hours",
                "median_ae_minutes",
            ]
        ]
        write_dataframe(ws, next_row + 2, 2, test_naive, "Naive baseline (test split)")


def build_experiment_summary_lines(exp_no: str, summary_df: pd.DataFrame) -> tuple[str, list[tuple[str, str]], list[str]]:
    if exp_no == "02":
        title = "Stage 3 baseline multitask: single-task vs multitask"
        rows = [
            ("모델 구조", "single-task baseline(anchor/refine) vs plain multitask"),
            ("비교 대상", "anchor_single_task, anchor_multi_task_w1.0, refine_single_task, refine_multi_task_w1.0"),
            ("main activity metric", "full ranking, NDCG@10"),
            ("main time metric", "MAE"),
        ]
        bullets = [
            f"anchor 기준으로 single-task({fnum(get_metric(summary_df, 'anchor_single_task', 'best_test_at_best_valid_full_ndcg@10', 'mean'))}) > multitask({fnum(get_metric(summary_df, 'anchor_multi_task', 'best_test_at_best_valid_full_ndcg@10', 'mean'))})",
            f"refine 기준으로도 single-task({fnum(get_metric(summary_df, 'refine_single_task', 'best_test_at_best_valid_full_ndcg@10', 'mean'))}) > multitask({fnum(get_metric(summary_df, 'refine_multi_task', 'best_test_at_best_valid_full_ndcg@10', 'mean'))})",
            "즉 baseline multitask 자체는 next activity 성능을 개선하지 못했고, multitask 내부에서는 anchor backbone이 refine보다 더 안정적이었음.",
        ]
        return title, rows, bullets

    if exp_no == "03":
        title = "Stage 3 attention-bias multitask (anchor backbone)"
        rows = [
            ("모델 구조", "anchor baseline / anchor attention bias / anchor multitask / anchor attnbias multitask"),
            ("비교 대상", "anchor_single_task, anchor_attnbias_single_task, anchor_multi_task_w1.0, anchor_attnbias_multi_task_w1.0"),
            ("time-aware", "delta_start + 9-bucket attention bias"),
            ("main activity metric", "full ranking, NDCG@10"),
        ]
        bullets = [
            f"anchor_attnbias_multi_task의 time MAE({fnum(get_metric(summary_df, 'anchor_attnbias_multi_task', 'best_test_at_best_valid_task_time_mae', 'mean'), 1)}초)는 plain multitask보다 좋았음.",
            f"하지만 next activity 기준으로는 attnbias multitask({fnum(get_metric(summary_df, 'anchor_attnbias_multi_task', 'best_test_at_best_valid_full_ndcg@10', 'mean'))}) < plain multitask({fnum(get_metric(summary_df, 'anchor_multi_task', 'best_test_at_best_valid_full_ndcg@10', 'mean'))}).",
            "즉 attention bias는 Stage 3에서 time 쪽은 일부 개선했지만, overall activity ranking은 더 악화시킴.",
        ]
        return title, rows, bullets

    if exp_no == "04":
        title = "Stage 3 attention-bias multitask: time_loss_weight 1.0 vs 0.1"
        rows = [
            ("모델 구조", "anchor attention-bias multitask"),
            ("비교 대상", "anchor_attnbias_multi_task_w1.0 vs anchor_attnbias_multi_task_w0.1"),
            ("조정 변수", "time_loss_weight"),
            ("main time metric", "MAE"),
        ]
        bullets = [
            f"w0.1 적용 시 next activity는 일부 회복({fnum(get_metric(summary_df, 'anchor_attnbias_multi_task_w0.1', 'best_test_at_best_valid_full_ndcg@10', 'mean'))} vs {fnum(get_metric(summary_df, 'anchor_attnbias_multi_task_w1.0', 'best_test_at_best_valid_full_ndcg@10', 'mean'))}).",
            f"반대로 time MAE는 악화({fnum(get_metric(summary_df, 'anchor_attnbias_multi_task_w0.1', 'best_test_at_best_valid_task_time_mae', 'mean'), 1)}초 vs {fnum(get_metric(summary_df, 'anchor_attnbias_multi_task_w1.0', 'best_test_at_best_valid_task_time_mae', 'mean'), 1)}초).",
            "즉 attention-bias multitask에서도 activity-time trade-off가 뚜렷하게 확인됨.",
        ]
        return title, rows, bullets

    if exp_no == "05":
        title = "Stage 3 plain anchor multitask: time_loss_weight 1.0 vs 0.1"
        rows = [
            ("모델 구조", "anchor baseline single-task vs anchor plain multitask"),
            ("비교 대상", "anchor_single_task, anchor_multi_task_w1.0, anchor_multi_task_w0.1"),
            ("조정 변수", "time_loss_weight"),
            ("main activity metric", "full ranking, NDCG@10"),
        ]
        bullets = [
            f"w0.1 적용으로 anchor multitask activity ranking이 크게 회복됨 ({fnum(get_metric(summary_df, 'anchor_multi_task_w0.1', 'best_test_at_best_valid_full_ndcg@10', 'mean'))} vs {fnum(get_metric(summary_df, 'anchor_multi_task_w1.0', 'best_test_at_best_valid_full_ndcg@10', 'mean'))}).",
            f"하지만 time MAE는 악화됨 ({fnum(get_metric(summary_df, 'anchor_multi_task_w0.1', 'best_test_at_best_valid_task_time_mae', 'mean'), 1)}초 vs {fnum(get_metric(summary_df, 'anchor_multi_task_w1.0', 'best_test_at_best_valid_task_time_mae', 'mean'), 1)}초).",
            "즉 plain multitask의 핵심 이슈는 backbone 자체보다도 loss balance에 가까웠고, Stage 3 best multitask(activity 기준)는 anchor_multi_task_w0.1임.",
        ]
        return title, rows, bullets

    if exp_no == "06":
        title = "Stage 3 plain refine multitask: time_loss_weight 1.0 vs 0.1"
        rows = [
            ("모델 구조", "refine baseline single-task vs refine plain multitask"),
            ("비교 대상", "refine_single_task, refine_multi_task_w1.0, refine_multi_task_w0.1"),
            ("조정 변수", "time_loss_weight"),
            ("main activity metric", "full ranking, NDCG@10"),
        ]
        bullets = [
            f"refine에서도 w0.1 적용으로 ranking이 회복됨 ({fnum(get_metric(summary_df, 'refine_multi_task_w0.1', 'best_test_at_best_valid_full_ndcg@10', 'mean'))} vs {fnum(get_metric(summary_df, 'refine_multi_task_w1.0', 'best_test_at_best_valid_full_ndcg@10', 'mean'))}).",
            f"time MAE는 소폭 악화({fnum(get_metric(summary_df, 'refine_multi_task_w0.1', 'best_test_at_best_valid_task_time_mae', 'mean'), 1)}초 vs {fnum(get_metric(summary_df, 'refine_multi_task_w1.0', 'best_test_at_best_valid_task_time_mae', 'mean'), 1)}초)했지만, anchor처럼 loss-balance 효과가 재현됨.",
            "다만 overall main metric 기준으로는 refine_single_task와 anchor_multi_task_w0.1를 넘지 못함.",
        ]
        return title, rows, bullets

    raise ValueError(f"Unsupported experiment no: {exp_no}")


def select_detail_columns(df: pd.DataFrame) -> pd.DataFrame:
    id_cols = [
        "run_name",
        "seed",
        "variant",
        "maxlen",
        "dropout_rate",
        "hidden_units",
        "selection_metric",
        "best_epoch",
        "enable_time_prediction",
        "time_prediction_target",
        "time_loss_weight",
        "time_target_transform",
        "time_modeling_mode",
    ]
    metric_prefixes = ("best_valid_", "best_test_at_best_valid_", "last_valid_", "last_test_")
    metric_cols = [c for c in df.columns if c.startswith(metric_prefixes)]
    metric_cols = ordered_metric_columns(metric_cols, RAW_METRIC_ORDER)
    cols = [c for c in id_cols if c in df.columns] + metric_cols
    return df[cols].copy()


def create_stage3_experiment_sheet(ws, template_ws, exp_no: str, parsed: NotebookTables) -> None:
    apply_template_styles(ws, template_ws)
    set_default_layout(ws)

    detail = select_detail_columns(clean_detail_df(parsed.detail))
    summary = clean_summary_df(parsed.summary)
    split_detail = build_split_raw_table(detail)
    split_summary = build_split_summary_table(summary, detail)

    title, summary_rows, bullets = build_experiment_summary_lines(exp_no, summary)
    next_row = write_summary_block(ws, title, summary_rows, bullets, max_col=34)

    if not split_summary.empty:
        next_row = write_stage2_like_summary_table(ws, next_row + 1, 2, split_summary, "전체 지표 mean/std summary")

    if not split_detail.empty:
        next_row = write_dataframe(ws, next_row + 2, 2, split_detail, "run별 raw 결과")


def create_stage3_naive_sheet(ws, template_ws, parsed07: NotebookTables) -> None:
    apply_template_styles(ws, template_ws)
    set_default_layout(ws)

    naive = clean_naive_df(parsed07.naive_readable)
    model_summary = clean_summary_df(parsed07.model_summary)

    rows = [
        ("실험 목적", "Stage 3 next-time 성능 해석을 위한 simple baseline 비교"),
        ("비교 대상", "global mean, global median, current activity mean, prefix length mean"),
        ("main time metric", "MAE"),
        ("해석 포인트", "trivial predictor와 process-aware simple baseline 대비 multitask time head 성능 확인"),
    ]
    bullets = [
        "global_mean / prefix_len_mean은 매우 약한 baseline으로 나타났고, 실제로 강한 simple baseline은 global_median과 activity_mean이었음.",
        "Stage 3 multitask들은 대체로 trivial baseline보다 낫지만, activity_mean 같은 process-aware simple baseline을 압도하진 못함.",
        "즉 next-time 예측은 의미는 있으나, simple heuristic 대비 아주 강하다고 보긴 어려움.",
    ]
    next_row = write_summary_block(ws, "Stage 3 next-time naive baseline 분석", rows, bullets, max_col=24)

    if not naive.empty:
        test_naive = naive[naive["split"] == "test"].copy()
        test_naive = test_naive[
            [
                "split",
                "baseline",
                "mae",
                "rmse",
                "median_ae",
                "mae_hours",
                "mae_minutes",
                "rmse_hours",
                "median_ae_minutes",
            ]
        ]
        write_dataframe(ws, next_row + 1, 2, test_naive, "Naive baseline (test split)")
        next_row += len(test_naive) + 4

    if not model_summary.empty:
        model_summary = model_summary.copy()
        cols = ["variant"]
        for metric in [
            "best_test_at_best_valid_task_time_mae",
            "best_test_at_best_valid_task_time_rmse",
            "best_test_at_best_valid_task_time_median_ae",
        ]:
            for stat in ("mean", "std"):
                key = f"{metric}|{stat}"
                if key in model_summary.columns:
                    cols.append(key)
        model_summary = model_summary[cols]
        write_dataframe(ws, next_row, 2, model_summary, "Stage 3 multitask time metric summary")


def ensure_backup(workbook_path: Path) -> Path:
    backup_path = workbook_path.with_name(f"{workbook_path.stem}__backup_before_stage3_{datetime.now():%Y%m%d_%H%M%S}{workbook_path.suffix}")
    copy2(workbook_path, backup_path)
    return backup_path


def main() -> None:
    parsed = {exp_no: load_notebook_tables(path) for exp_no, path in STAGE3_NOTEBOOKS.items()}

    backup_path = ensure_backup(WORKBOOK_PATH)
    wb = openpyxl.load_workbook(WORKBOOK_PATH)

    for name in NEW_SHEETS:
        if name in wb.sheetnames:
            del wb[name]

    template_ws = wb["2-2.attnbias_refine"]

    overview_ws = wb.create_sheet("3-0.Stage3")
    create_stage3_overview(overview_ws, template_ws, parsed)

    sheet_map = {
        "3-1.Baseline_mt": "02",
        "3-2.AttnBias_mt": "03",
        "3-3.AttnBias_w01": "04",
        "3-4.Anchor_w01": "05",
        "3-5.Refine_w01": "06",
    }
    for sheet_name, exp_no in sheet_map.items():
        ws = wb.create_sheet(sheet_name)
        create_stage3_experiment_sheet(ws, template_ws, exp_no, parsed[exp_no])

    naive_ws = wb.create_sheet("3-6.Time_naive")
    create_stage3_naive_sheet(naive_ws, template_ws, parsed["07"])

    wb.save(WORKBOOK_PATH)
    print(f"[ok] backup created: {backup_path}")
    print(f"[ok] workbook updated: {WORKBOOK_PATH}")
    print("[ok] added sheets:", ", ".join(NEW_SHEETS))


if __name__ == "__main__":
    main()
