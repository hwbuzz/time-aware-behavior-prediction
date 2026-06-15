from __future__ import annotations

from html.parser import HTMLParser
from pathlib import Path

import nbformat
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
NOTEBOOK_PATH = ROOT / "notebooks" / "sasrec_stage3_bpi2012_colab_train_10_260614.ipynb"
WORKBOOK_PATH = ROOT / "docs" / "임시_stage2_sinusoidal_added_20260614_132150.xlsx"
SHEET_NAME = "3-7.Sinusoidal_MT"

DETAIL_METRICS = [
    "best_valid_full_ndcg@10",
    "best_valid_full_hr@10",
    "best_valid_full_ndcg@5",
    "best_valid_full_hr@5",
    "best_valid_full_mrr",
    "best_test_at_best_valid_full_ndcg@10",
    "best_test_at_best_valid_full_hr@10",
    "best_test_at_best_valid_full_ndcg@5",
    "best_test_at_best_valid_full_hr@5",
    "best_test_at_best_valid_full_mrr",
    "best_valid_sampled_ndcg@10",
    "best_valid_sampled_hr@10",
    "best_valid_sampled_ndcg@5",
    "best_valid_sampled_hr@5",
    "best_valid_sampled_mrr",
    "best_test_at_best_valid_sampled_ndcg@10",
    "best_test_at_best_valid_sampled_hr@10",
    "best_test_at_best_valid_sampled_ndcg@5",
    "best_test_at_best_valid_sampled_hr@5",
    "best_test_at_best_valid_sampled_mrr",
    "best_valid_task_accuracy",
    "best_valid_task_macro_f1",
    "best_valid_task_top5_accuracy",
    "best_valid_task_top10_accuracy",
    "best_valid_task_time_mae",
    "best_valid_task_time_rmse",
    "best_valid_task_time_median_ae",
    "best_test_at_best_valid_task_accuracy",
    "best_test_at_best_valid_task_macro_f1",
    "best_test_at_best_valid_task_top5_accuracy",
    "best_test_at_best_valid_task_top10_accuracy",
    "best_test_at_best_valid_task_time_mae",
    "best_test_at_best_valid_task_time_rmse",
    "best_test_at_best_valid_task_time_median_ae",
]

SUMMARY_VARIANTS = [
    "refine_baseline",
    "refine_attnbias_single_task",
    "refine_sinusoidal_single_task",
    "refine_multi_task_w1.0",
    "refine_attnbias_multi_task_w1.0",
    "refine_sinusoidal_multi_task_w1.0",
    "refine_sinusoidal_multi_task_w0.1",
]


class SimpleTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._current_row: list[str] = []
        self._in_cell = False
        self._buf = ""

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self._current_row = []
        elif tag in ("th", "td"):
            self._in_cell = True
            self._buf = ""

    def handle_endtag(self, tag):
        if tag in ("th", "td") and self._in_cell:
            self._current_row.append(self._buf.strip())
            self._in_cell = False
        elif tag == "tr" and self._current_row:
            self.rows.append(self._current_row)

    def handle_data(self, data):
        if self._in_cell:
            self._buf += data


def convert_value(value):
    if value in ("", "None", "nan", "NaN", None):
        return None
    if isinstance(value, str):
        low = value.lower()
        if low == "true":
            return True
        if low == "false":
            return False
        try:
            if "." not in value and "e" not in low:
                return int(value)
            return float(value)
        except ValueError:
            return value
    return value


def notebook_df() -> pd.DataFrame:
    nb = nbformat.read(NOTEBOOK_PATH, as_version=4)
    html = nb.cells[35]["outputs"][0]["data"]["text/html"]
    p = SimpleTableParser()
    p.feed(html)
    header = p.rows[0][1:]
    rows = []
    for row in p.rows[1:]:
        vals = row[1:]
        if len(vals) == len(header):
            rows.append({k: convert_value(v) for k, v in zip(header, vals)})
    return pd.DataFrame(rows).sort_values(["variant", "seed", "run_name"]).reset_index(drop=True)


def workbook_detail_df() -> pd.DataFrame:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 5).value == "run_name":
            header_row = r
            break
    if header_row is None:
        raise ValueError("detail header row not found")
    rows = []
    r = header_row + 1
    while r <= ws.max_row:
        run_name = ws.cell(r, 5).value
        seq = ws.cell(r, 2).value
        if run_name is None and seq is None:
            break
        rows.append({
            "run_name": run_name,
            "seed": ws.cell(r, 6).value,
            "variant": ws.cell(r, 4).value,
            "maxlen": ws.cell(r, 7).value,
            "dropout_rate": ws.cell(r, 8).value,
            "time_encoding": ws.cell(r, 9).value,
            "time_delta_column": ws.cell(r, 10).value,
            "time_loss_weight": ws.cell(r, 11).value,
            "eval_type": ws.cell(r, 12).value,
            **{metric: ws.cell(r, 14 + idx).value for idx, metric in enumerate(DETAIL_METRICS)}
        })
        r += 1
    return pd.DataFrame(rows).sort_values(["variant", "seed", "run_name"]).reset_index(drop=True)


def workbook_summary_df() -> pd.DataFrame:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    section_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 2).value == "▶ run별 성능 평균, 표준편차":
            section_row = r
            break
    if section_row is None:
        raise ValueError("summary section not found")
    data_start = section_row + 4
    rows = []
    r = data_start
    while r <= ws.max_row:
        variant = ws.cell(r, 4).value
        if variant is None:
            break
        row = {"variant": variant}
        row["eval_type"] = ws.cell(r, 8).value
        c = 10
        for metric in DETAIL_METRICS:
            row[f"{metric}__mean"] = ws.cell(r, c).value
            row[f"{metric}__std"] = ws.cell(r, c + 1).value
            c += 2
        rows.append(row)
        r += 1
    return pd.DataFrame(rows)


def same(a, b, tol=1e-9):
    if pd.isna(a) and pd.isna(b):
        return True
    if a is None and b is None:
        return True
    try:
        if isinstance(a, (int, float)) or isinstance(b, (int, float)):
            return abs(float(a) - float(b)) <= tol
    except Exception:
        pass
    return a == b


def main() -> None:
    ndf = notebook_df()
    wdf = workbook_detail_df()
    detail_mismatches = []
    expanded_rows = []
    for _, row in ndf.iterrows():
        base = {
            "run_name": row["run_name"],
            "seed": row["seed"],
            "variant": row["variant"],
            "maxlen": row["maxlen"],
            "dropout_rate": row["dropout_rate"],
            "time_encoding": row["time_encoding"],
            "time_delta_column": row["time_delta_column"],
            "time_loss_weight": row["time_loss_weight"],
        }
        full = base | {
            "eval_type": "full ranking",
            "best_valid_full_ndcg@10": row["best_valid_full_ndcg@10"],
            "best_valid_full_hr@10": row["best_valid_full_hr@10"],
            "best_valid_full_ndcg@5": row["best_valid_full_ndcg@5"],
            "best_valid_full_hr@5": row["best_valid_full_hr@5"],
            "best_valid_full_mrr": row["best_valid_full_mrr"],
            "best_test_at_best_valid_full_ndcg@10": row["best_test_at_best_valid_full_ndcg@10"],
            "best_test_at_best_valid_full_hr@10": row["best_test_at_best_valid_full_hr@10"],
            "best_test_at_best_valid_full_ndcg@5": row["best_test_at_best_valid_full_ndcg@5"],
            "best_test_at_best_valid_full_hr@5": row["best_test_at_best_valid_full_hr@5"],
            "best_test_at_best_valid_full_mrr": row["best_test_at_best_valid_full_mrr"],
            "best_valid_sampled_ndcg@5": None,
            "best_valid_sampled_hr@5": None,
            "best_valid_sampled_ndcg@10": None,
            "best_valid_sampled_hr@10": None,
            "best_valid_sampled_mrr": None,
            "best_test_at_best_valid_sampled_ndcg@5": None,
            "best_test_at_best_valid_sampled_hr@5": None,
            "best_test_at_best_valid_sampled_ndcg@10": None,
            "best_test_at_best_valid_sampled_hr@10": None,
            "best_test_at_best_valid_sampled_mrr": None,
        }
        sampled = base | {
            "eval_type": "negative sampling(100)",
            "best_valid_full_ndcg@10": row["best_valid_sampled_ndcg@10"],
            "best_valid_full_hr@10": row["best_valid_sampled_hr@10"],
            "best_valid_full_ndcg@5": row["best_valid_sampled_ndcg@5"],
            "best_valid_full_hr@5": row["best_valid_sampled_hr@5"],
            "best_valid_full_mrr": row["best_valid_sampled_mrr"],
            "best_test_at_best_valid_full_ndcg@10": row["best_test_at_best_valid_sampled_ndcg@10"],
            "best_test_at_best_valid_full_hr@10": row["best_test_at_best_valid_sampled_hr@10"],
            "best_test_at_best_valid_full_ndcg@5": row["best_test_at_best_valid_sampled_ndcg@5"],
            "best_test_at_best_valid_full_hr@5": row["best_test_at_best_valid_sampled_hr@5"],
            "best_test_at_best_valid_full_mrr": row["best_test_at_best_valid_sampled_mrr"],
            "best_valid_sampled_ndcg@5": None,
            "best_valid_sampled_hr@5": None,
            "best_valid_sampled_ndcg@10": None,
            "best_valid_sampled_hr@10": None,
            "best_valid_sampled_mrr": None,
            "best_test_at_best_valid_sampled_ndcg@5": None,
            "best_test_at_best_valid_sampled_hr@5": None,
            "best_test_at_best_valid_sampled_ndcg@10": None,
            "best_test_at_best_valid_sampled_hr@10": None,
            "best_test_at_best_valid_sampled_mrr": None,
        }
        for task_metric in [
            "best_valid_task_accuracy","best_valid_task_macro_f1","best_valid_task_top5_accuracy","best_valid_task_top10_accuracy",
            "best_valid_task_time_mae","best_valid_task_time_rmse","best_valid_task_time_median_ae",
            "best_test_at_best_valid_task_accuracy","best_test_at_best_valid_task_macro_f1","best_test_at_best_valid_task_top5_accuracy","best_test_at_best_valid_task_top10_accuracy",
            "best_test_at_best_valid_task_time_mae","best_test_at_best_valid_task_time_rmse","best_test_at_best_valid_task_time_median_ae",
        ]:
            full[task_metric] = row[task_metric]
            sampled[task_metric] = row[task_metric]
        expanded_rows.extend([full, sampled])
    ndf2 = pd.DataFrame(expanded_rows).sort_values(["variant", "seed", "run_name", "eval_type"]).reset_index(drop=True)
    wdf = wdf.sort_values(["variant", "seed", "run_name", "eval_type"]).reset_index(drop=True)
    compare_cols = ["run_name", "seed", "variant", "maxlen", "dropout_rate", "time_encoding", "time_delta_column", "time_loss_weight", "eval_type"] + [
        "best_valid_full_ndcg@10","best_valid_full_hr@10","best_valid_full_ndcg@5","best_valid_full_hr@5","best_valid_full_mrr",
        "best_test_at_best_valid_full_ndcg@10","best_test_at_best_valid_full_hr@10","best_test_at_best_valid_full_ndcg@5","best_test_at_best_valid_full_hr@5","best_test_at_best_valid_full_mrr",
        "best_valid_task_accuracy","best_valid_task_macro_f1","best_valid_task_top5_accuracy","best_valid_task_top10_accuracy","best_valid_task_time_mae","best_valid_task_time_rmse","best_valid_task_time_median_ae",
        "best_test_at_best_valid_task_accuracy","best_test_at_best_valid_task_macro_f1","best_test_at_best_valid_task_top5_accuracy","best_test_at_best_valid_task_top10_accuracy","best_test_at_best_valid_task_time_mae","best_test_at_best_valid_task_time_rmse","best_test_at_best_valid_task_time_median_ae",
    ]
    if len(ndf2) != len(wdf):
        detail_mismatches.append(("row_count", len(ndf2), len(wdf)))
    else:
        for i in range(len(ndf2)):
            for col in compare_cols:
                if not same(ndf2.loc[i, col], wdf.loc[i, col]):
                    detail_mismatches.append((i, col, ndf2.loc[i, col], wdf.loc[i, col]))

    expected_summary = ndf.groupby("variant")[DETAIL_METRICS].agg(["mean", "std"]).reindex(SUMMARY_VARIANTS)
    actual_summary = workbook_summary_df()
    summary_mismatches = []
    for variant in SUMMARY_VARIANTS:
        full_row = actual_summary[(actual_summary["variant"] == variant) & (actual_summary["eval_type"] == "full ranking")]
        sampled_row = actual_summary[(actual_summary["variant"] == variant) & (actual_summary["eval_type"] == "negative sampling(100)")]
        if len(full_row) != 1 or len(sampled_row) != 1:
            summary_mismatches.append((variant, "row_count", len(full_row), len(sampled_row)))
            continue
        full_row = full_row.iloc[0]
        sampled_row = sampled_row.iloc[0]
        for metric in DETAIL_METRICS:
            target_row = full_row
            if "_sampled_" in metric:
                target_row = sampled_row
            a = expected_summary.loc[variant, (metric, "mean")]
            b = target_row[f"{metric}__mean"]
            if not same(a, b):
                summary_mismatches.append((variant, metric, "mean", a, b))
            a = expected_summary.loc[variant, (metric, "std")]
            b = target_row[f"{metric}__std"]
            if not same(a, b):
                summary_mismatches.append((variant, metric, "std", a, b))

    print("detail row counts:", len(ndf2), len(wdf))
    print("detail mismatches:", len(detail_mismatches))
    if detail_mismatches:
        for item in detail_mismatches[:20]:
            print("DETAIL_MISMATCH", item)
    print("summary mismatches:", len(summary_mismatches))
    if summary_mismatches:
        for item in summary_mismatches[:20]:
            print("SUMMARY_MISMATCH", item)
    if not detail_mismatches and not summary_mismatches:
        print("[ok] workbook values match notebook outputs")


if __name__ == "__main__":
    main()
