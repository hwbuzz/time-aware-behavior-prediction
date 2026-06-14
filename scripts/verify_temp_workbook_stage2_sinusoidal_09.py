from __future__ import annotations

from html.parser import HTMLParser
from pathlib import Path

import nbformat
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
NOTEBOOK_PATH = ROOT / "notebooks" / "sasrec_timeaware_bpi2012_colab_train_09_260612.ipynb"
WORKBOOK_PATH = ROOT / "docs" / "임시_stage2_sinusoidal_added_20260614_132150.xlsx"
SHEET_NAME = "2-3.Sinusoidal"


DETAIL_METRICS = [
    "best_valid_full_ndcg@10",
    "best_valid_full_hr@10",
    "best_valid_full_ndcg@5",
    "best_valid_full_hr@5",
    "best_valid_full_mrr",
    "best_test_at_best_valid_full_ndcg@10",
    "best_test_at_best_valid_full_hr@10",
    "best_test_at_best_valid_full_ndcg@5",
    "best_test_at_best_valid_full_hr@5",
    "best_test_at_best_valid_full_mrr",
    "best_valid_sampled_ndcg@10",
    "best_valid_sampled_hr@10",
    "best_valid_sampled_ndcg@5",
    "best_valid_sampled_hr@5",
    "best_valid_sampled_mrr",
    "best_test_at_best_valid_sampled_ndcg@10",
    "best_test_at_best_valid_sampled_hr@10",
    "best_test_at_best_valid_sampled_ndcg@5",
    "best_test_at_best_valid_sampled_hr@5",
    "best_test_at_best_valid_sampled_mrr",
    "best_valid_task_accuracy",
    "best_valid_task_macro_f1",
    "best_valid_task_top5_accuracy",
    "best_valid_task_top10_accuracy",
    "best_test_at_best_valid_task_accuracy",
    "best_test_at_best_valid_task_macro_f1",
    "best_test_at_best_valid_task_top5_accuracy",
    "best_test_at_best_valid_task_top10_accuracy",
]


SUMMARY_VARIANTS = [
    "anchor_baseline",
    "anchor_bucket_b9",
    "anchor_attnbias_dstart_b9",
    "anchor_sinusoidal_dprev",
    "anchor_sinusoidal_dstart",
    "refine_baseline",
    "refine_bucket_b9",
    "refine_dstart_bucket_b9",
    "refine_dstart_continuous",
    "refine_attnbias_dstart_b9",
    "refine_sinusoidal_dprev",
    "refine_sinusoidal_dstart",
]


class SimpleTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._current_row: list[str] = []
        self._in_cell = False
        self._buf = ""

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self._current_row = []
        elif tag in ("th", "td"):
            self._in_cell = True
            self._buf = ""

    def handle_endtag(self, tag):
        if tag in ("th", "td") and self._in_cell:
            self._current_row.append(self._buf.strip())
            self._in_cell = False
        elif tag == "tr" and self._current_row:
            self.rows.append(self._current_row)

    def handle_data(self, data):
        if self._in_cell:
            self._buf += data


def convert_value(value):
    if value in ("", "None", "nan", "NaN", None):
        return None
    if isinstance(value, str):
        low = value.lower()
        if low == "true":
            return True
        if low == "false":
            return False
        try:
            if "." not in value and "e" not in low:
                return int(value)
            return float(value)
        except ValueError:
            return value
    return value


def parse_run_table_from_notebook() -> pd.DataFrame:
    nb = nbformat.read(NOTEBOOK_PATH, as_version=4)
    html = nb.cells[44]["outputs"][0]["data"]["text/html"]
    parser = SimpleTableParser()
    parser.feed(html)
    header = parser.rows[0][1:]
    records = []
    for row in parser.rows[1:]:
        values = row[1:]
        if len(values) != len(header):
            continue
        records.append({k: convert_value(v) for k, v in zip(header, values)})
    return pd.DataFrame(records)


def parse_detail_table_from_workbook() -> pd.DataFrame:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # find detail header row by matching run_name header
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 5).value == "run_name":
            header_row = r
            break
    if header_row is None:
        raise ValueError("detail header row not found")

    data_start = header_row + 1
    rows = []
    r = data_start
    while r <= ws.max_row:
        seq = ws.cell(r, 2).value
        run_name = ws.cell(r, 5).value
        if seq is None and run_name is None:
            break
        row = {
            "run_name": run_name,
            "seed": ws.cell(r, 6).value,
            "variant": ws.cell(r, 4).value,
            "maxlen": ws.cell(r, 7).value,
            "dropout_rate": ws.cell(r, 8).value,
            "time_encoding": ws.cell(r, 9).value,
            "time_delta_column": ws.cell(r, 10).value,
            "selection_metric": "full_valid_ndcg@10",
        }
        for idx, metric in enumerate(DETAIL_METRICS, start=12):
            row[metric] = ws.cell(r, idx).value
        rows.append(row)
        r += 1
    return pd.DataFrame(rows)


def parse_summary_table_from_workbook() -> pd.DataFrame:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # find mean/std section by section title
    section_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 2).value == "▶ run별 성능 평균, 표준편차":
            section_row = r
            break
    if section_row is None:
        raise ValueError("mean/std section not found")

    header_row = section_row + 3
    data_start = section_row + 4

    rows = []
    r = data_start
    while r <= ws.max_row:
        variant = ws.cell(r, 4).value
        if variant is None:
            break
        row = {
            "variant": variant,
            "maxlen": ws.cell(r, 5).value,
            "dropout_rate": ws.cell(r, 6).value,
        }
        c = 8
        for metric in DETAIL_METRICS:
            row[f"{metric}__mean"] = ws.cell(r, c).value
            row[f"{metric}__std"] = ws.cell(r, c + 1).value
            c += 2
        rows.append(row)
        r += 1
    return pd.DataFrame(rows)


def compare_values(a, b, tol=1e-9) -> bool:
    if pd.isna(a) and pd.isna(b):
        return True
    if a is None and b is None:
        return True
    if isinstance(a, (int, float)) or isinstance(b, (int, float)):
        try:
            return abs(float(a) - float(b)) <= tol
        except Exception:
            return a == b
    return a == b


def main() -> None:
    notebook_df = parse_run_table_from_notebook().sort_values(["variant", "seed", "run_name"]).reset_index(drop=True)
    workbook_detail_df = parse_detail_table_from_workbook().sort_values(["variant", "seed", "run_name"]).reset_index(drop=True)

    print("detail row counts:", len(notebook_df), len(workbook_detail_df))
    detail_mismatches = []
    if len(notebook_df) != len(workbook_detail_df):
        detail_mismatches.append(("row_count", len(notebook_df), len(workbook_detail_df)))
    else:
        compare_cols = [
            "run_name", "seed", "variant", "maxlen", "dropout_rate", "time_encoding", "time_delta_column"
        ] + DETAIL_METRICS
        for idx in range(len(notebook_df)):
            for col in compare_cols:
                a = notebook_df.loc[idx, col]
                b = workbook_detail_df.loc[idx, col]
                if not compare_values(a, b):
                    detail_mismatches.append((idx, col, a, b))

    summary_expected = (
        notebook_df.groupby("variant")[DETAIL_METRICS].agg(["mean", "std"]).reindex(SUMMARY_VARIANTS)
    )
    workbook_summary_df = parse_summary_table_from_workbook().set_index("variant").reindex(SUMMARY_VARIANTS)

    summary_mismatches = []
    for variant in SUMMARY_VARIANTS:
        for metric in DETAIL_METRICS:
            for stat in ("mean", "std"):
                a = summary_expected.loc[variant, (metric, stat)]
                b = workbook_summary_df.loc[variant, f"{metric}__{stat}"]
                if not compare_values(a, b):
                    summary_mismatches.append((variant, metric, stat, a, b))

    print("detail mismatches:", len(detail_mismatches))
    if detail_mismatches:
        for item in detail_mismatches[:20]:
            print("DETAIL_MISMATCH", item)

    print("summary mismatches:", len(summary_mismatches))
    if summary_mismatches:
        for item in summary_mismatches[:20]:
            print("SUMMARY_MISMATCH", item)

    if not detail_mismatches and not summary_mismatches:
        print("[ok] workbook values match notebook outputs")


if __name__ == "__main__":
    main()
