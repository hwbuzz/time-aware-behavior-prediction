from __future__ import annotations

from datetime import datetime
from pathlib import Path
import shutil
import sys

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.add_stage3_sheets_to_capstone_workbook import (
    apply_template_styles,
    build_split_raw_table,
    build_split_summary_table,
    clean_detail_df,
    clean_summary_df,
    fnum,
    get_metric,
    load_notebook_tables,
    select_detail_columns,
    set_default_layout,
    write_dataframe,
    write_stage2_like_summary_table,
    write_summary_block,
)


WORKBOOK_PATH = Path("docs/임시.xlsx")
NOTEBOOK_PATH = Path("notebooks/sasrec_stage3_bpi2012_colab_train_08_260607.ipynb")
TARGET_SHEET = "3-8.RefineAttnBias_mt"


def build_summary_lines(summary_df):
    title = "Stage 3 refine attention-bias multitask: w1.0 vs w0.1"
    rows = [
        ("모델 구조", "refine baseline / refine attention bias / refine plain multitask / refine attnbias multitask"),
        ("비교 대상", "refine_single_task, refine_attnbias_single_task, refine_multi_task_w1.0, refine_attnbias_multi_task_w1.0, refine_attnbias_multi_task_w0.1"),
        ("time-aware", "delta_start + 9-bucket attention bias"),
        ("main activity metric", "full ranking, NDCG@10"),
        ("main time metric", "MAE"),
    ]
    bullets = [
        (
            f"refine 기준 single-task baseline({fnum(get_metric(summary_df, 'refine_single_task', 'best_test_at_best_valid_full_ndcg@10', 'mean'))})이 "
            f"attention-bias single-task({fnum(get_metric(summary_df, 'refine_attnbias_single_task', 'best_test_at_best_valid_full_ndcg@10', 'mean'))})보다 높음."
        ),
        (
            f"plain refine multitask({fnum(get_metric(summary_df, 'refine_multi_task_w1.0', 'best_test_at_best_valid_full_ndcg@10', 'mean'))}) 대비 "
            f"attnbias multitask w1.0({fnum(get_metric(summary_df, 'refine_attnbias_multi_task_w1.0', 'best_test_at_best_valid_full_ndcg@10', 'mean'))})는 activity ranking이 더 낮음."
        ),
        (
            f"w0.1 적용 시 activity는 일부 회복({fnum(get_metric(summary_df, 'refine_attnbias_multi_task_w0.1', 'best_test_at_best_valid_full_ndcg@10', 'mean'))})되지만, "
            f"time MAE는 {fnum(get_metric(summary_df, 'refine_attnbias_multi_task_w0.1', 'best_test_at_best_valid_task_time_mae', 'mean'), 1)}초로 "
            f"w1.0({fnum(get_metric(summary_df, 'refine_attnbias_multi_task_w1.0', 'best_test_at_best_valid_task_time_mae', 'mean'), 1)}초) 대비 trade-off가 나타남."
        ),
    ]
    return title, rows, bullets


def main():
    backup_path = WORKBOOK_PATH.with_name(
        f"{WORKBOOK_PATH.stem}__backup_before_08_{datetime.now().strftime('%Y%m%d_%H%M%S')}{WORKBOOK_PATH.suffix}"
    )
    shutil.copy2(WORKBOOK_PATH, backup_path)

    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    template_ws = wb[wb.sheetnames[0]]

    parsed = load_notebook_tables(NOTEBOOK_PATH)
    detail = select_detail_columns(clean_detail_df(parsed.detail))
    summary = clean_summary_df(parsed.summary)
    split_detail = build_split_raw_table(detail)
    split_summary = build_split_summary_table(summary, detail)
    title, summary_rows, bullets = build_summary_lines(summary)

    ws = wb.create_sheet(TARGET_SHEET)
    apply_template_styles(ws, template_ws)
    set_default_layout(ws)

    next_row = write_summary_block(ws, title, summary_rows, bullets, max_col=34)
    if not split_summary.empty:
        next_row = write_stage2_like_summary_table(ws, next_row + 1, 2, split_summary, "전체 지표 mean/std summary")
    if not split_detail.empty:
        write_dataframe(ws, next_row + 2, 2, split_detail, "run별 raw 결과")

    wb.remove(template_ws)
    wb.save(WORKBOOK_PATH)
    print(f"[ok] backup created: {backup_path}")
    print(f"[ok] workbook updated: {WORKBOOK_PATH}")
    print(f"[ok] sheet written: {TARGET_SHEET}")


if __name__ == "__main__":
    main()
