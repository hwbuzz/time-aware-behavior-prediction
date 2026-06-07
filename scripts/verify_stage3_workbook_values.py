from __future__ import annotations

import math
from pathlib import Path

import openpyxl

from scripts.add_stage3_sheets_to_capstone_workbook import (
    SUMMARY_METRIC_LAYOUT,
    STAGE3_NOTEBOOKS,
    build_overview_df,
    build_split_raw_table,
    build_split_summary_table,
    clean_detail_df,
    clean_naive_df,
    clean_summary_df,
    load_notebook_tables,
)


WORKBOOK_PATH = Path("docs/캡스톤_모델성능_v0.5_260604.xlsx")


def find_row(ws, text: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 2).value == text:
            return r
    return None


def cell_eq(a, b, tol: float = 1e-9) -> bool:
    if a is None and (b is None or b == ""):
        return True
    if b is None and (a is None or a == ""):
        return True
    try:
        fa = float(a)
        fb = float(b)
        return math.isclose(fa, fb, rel_tol=0.0, abs_tol=tol)
    except Exception:
        return str(a) == str(b)


def check_df(ws, title: str, expected):
    row = find_row(ws, title)
    if row is None:
        return [f"missing title row: {title}"]
    header_row = row + 1
    data_row = row + 2
    errs: list[str] = []

    actual_headers = [ws.cell(header_row, 2 + i).value for i in range(len(expected.columns))]
    expected_headers = list(expected.columns)
    if actual_headers != expected_headers:
        errs.append(f"{title}: header mismatch")
        for idx, (a, e) in enumerate(zip(actual_headers, expected_headers), start=1):
            if a != e:
                errs.append(f"  col {idx}: actual={a!r} expected={e!r}")
        if len(actual_headers) != len(expected_headers):
            errs.append(f"  header len actual={len(actual_headers)} expected={len(expected_headers)}")
        return errs

    for r_idx in range(len(expected)):
        for c_idx, col in enumerate(expected.columns):
            actual = ws.cell(data_row + r_idx, 2 + c_idx).value
            exp = expected.iloc[r_idx, c_idx]
            if isinstance(exp, float) and math.isnan(exp):
                exp = None
            if not cell_eq(actual, exp):
                errs.append(
                    f"{title}: row {r_idx + 1} col {col} actual={actual!r} expected={exp!r}"
                )
                if len(errs) >= 20:
                    return errs
    return errs


def check_stage2_like_summary_df(ws, title: str, expected):
    row = find_row(ws, title)
    if row is None:
        return [f"missing title row: {title}"]

    data_row = row + 4
    errs: list[str] = []
    meta_headers = ["SEQ", "Dataset", "variant", "maxlen", "dropout_rate", "time_loss_weight", "평가 방식", "best epoch 기준"]

    for r_idx in range(len(expected)):
        excel_row = data_row + r_idx

        for c_idx, col in enumerate(meta_headers):
            actual = ws.cell(excel_row, 2 + c_idx).value
            exp = expected.iloc[r_idx][col] if col in expected.columns else None
            if isinstance(exp, float) and math.isnan(exp):
                exp = None
            if not cell_eq(actual, exp):
                errs.append(f"{title}: row {r_idx + 1} meta {col} actual={actual!r} expected={exp!r}")
                if len(errs) >= 20:
                    return errs

        valid_start = 2 + len(meta_headers)
        current_col = valid_start
        row_data = expected.iloc[r_idx]
        for _, v_mean, v_std, _, _ in SUMMARY_METRIC_LAYOUT:
            for col_name, offset in ((v_mean, 0), (v_std, 1)):
                actual = ws.cell(excel_row, current_col + offset).value
                exp = row_data[col_name] if col_name in expected.columns else None
                if isinstance(exp, float) and math.isnan(exp):
                    exp = None
                if not cell_eq(actual, exp):
                    errs.append(
                        f"{title}: row {r_idx + 1} col {col_name} actual={actual!r} expected={exp!r}"
                    )
                    if len(errs) >= 20:
                        return errs
            current_col += 2

        test_start = valid_start + (len(SUMMARY_METRIC_LAYOUT) * 2)
        current_col = test_start
        for _, _, _, t_mean, t_std in SUMMARY_METRIC_LAYOUT:
            for col_name, offset in ((t_mean, 0), (t_std, 1)):
                actual = ws.cell(excel_row, current_col + offset).value
                exp = row_data[col_name] if col_name in expected.columns else None
                if isinstance(exp, float) and math.isnan(exp):
                    exp = None
                if not cell_eq(actual, exp):
                    errs.append(
                        f"{title}: row {r_idx + 1} col {col_name} actual={actual!r} expected={exp!r}"
                    )
                    if len(errs) >= 20:
                        return errs
            current_col += 2
    return errs


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    parsed = {key: load_notebook_tables(path) for key, path in STAGE3_NOTEBOOKS.items()}

    sheet_map = {
        "3-1.Baseline_mt": "02",
        "3-2.AttnBias_mt": "03",
        "3-3.AttnBias_w01": "04",
        "3-4.Anchor_w01": "05",
        "3-5.Refine_w01": "06",
    }

    mismatches: list[tuple[str, list[str]]] = []

    for sheet_name, exp_no in sheet_map.items():
        ws = wb[sheet_name]
        detail = clean_detail_df(parsed[exp_no].detail)
        summary = clean_summary_df(parsed[exp_no].summary)
        exp_summary = build_split_summary_table(summary, detail)
        exp_detail = build_split_raw_table(detail)
        errs = []
        errs += check_stage2_like_summary_df(ws, "전체 지표 mean/std summary", exp_summary)
        errs += check_df(ws, "run별 raw 결과", exp_detail)
        if errs:
            mismatches.append((sheet_name, errs))

    # 3-0 overview
    ws = wb["3-0.Stage3"]
    exp_overview = build_overview_df(parsed)
    errs = check_df(ws, "핵심 variant 비교", exp_overview)
    if errs:
        mismatches.append(("3-0.Stage3 overview", errs))

    # 3-6 naive baselines
    ws = wb["3-6.Time_naive"]
    naive = clean_naive_df(parsed["07"].naive_readable)
    exp_naive = naive[naive["split"] == "test"].copy()[
        [
            "split",
            "baseline",
            "mae",
            "rmse",
            "median_ae",
            "mae_hours",
            "mae_minutes",
            "rmse_hours",
            "median_ae_minutes",
        ]
    ]
    errs = check_df(ws, "Naive baseline (test split)", exp_naive)
    if errs:
        mismatches.append(("3-6.Time_naive naive", errs))

    model_summary = clean_summary_df(parsed["07"].model_summary)
    cols = ["variant"]
    for metric in [
        "best_test_at_best_valid_task_time_mae",
        "best_test_at_best_valid_task_time_rmse",
        "best_test_at_best_valid_task_time_median_ae",
    ]:
        for stat in ("mean", "std"):
            key = f"{metric}|{stat}"
            if key in model_summary.columns:
                cols.append(key)
    exp_model = model_summary[cols]
    errs = check_df(ws, "Stage 3 multitask time metric summary", exp_model)
    if errs:
        mismatches.append(("3-6.Time_naive model summary", errs))

    if not mismatches:
        print("ALL_OK")
        return

    print("MISMATCHES_FOUND")
    for sheet_name, errs in mismatches:
        print(f"\n## {sheet_name}")
        for err in errs:
            print(err)


if __name__ == "__main__":
    main()
